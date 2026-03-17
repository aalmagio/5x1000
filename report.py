#!/usr/bin/env python3
"""
5 per Mille - Generatore report Excel stile ASSIF/Bedogni.

Combina i dati delle 7 categorie in un workbook multi-foglio con:
- COMPLESSIVO: tutte le entita' con breakdown per categoria
- 7 fogli per categoria: dati filtrati con classifica interna
- Grafici: aggregazione regionale
- Licenza: attribuzione dati

Uso:
    python report.py --anno 2024
    python report.py --anno 2024 --no-confronto
    python report.py --anno 2024 --anno-confronto 2022
    python report.py --anno 2024 --output mio_report.xlsx
    python report.py --help

Requisiti:
    pip install openpyxl xlsxwriter
"""

import sys
import os
import re
import glob
import logging
import argparse
from pathlib import Path

import openpyxl
import xlsxwriter

# Importa utility dal progetto
from etl import (
    parse_importo,
    parse_intero,
    normalize_cf,
    EXCEL_CONFIG,
    _load_config,
    _apply_config,
)
from cinque_per_mille import setup_logging


# ============================================================================
# COSTANTI
# ============================================================================

CATEGORIES = [
    {
        "key": "volontariato",
        "display": "Enti del\nVOLONTARIATO",
        "display_short": "Volontariato",
        "sheet": "Volontariato",
        "slug": "ETS_ONLUS",
        "has_redistribuzione": True,
    },
    {
        "key": "sport",
        "display": "Enti dello\nSPORT",
        "display_short": "Sport",
        "sheet": "Sport",
        "slug": "ASD",
        "has_redistribuzione": True,
    },
    {
        "key": "ricerca_sci",
        "display": "Enti della\nRICERCA\nSCIENTIFICA",
        "display_short": "Ric. Scientifica",
        "sheet": "Scienza",
        "slug": "Ricerca_scientifica",
        "has_redistribuzione": True,
    },
    {
        "key": "ricerca_san",
        "display": "Enti della\nRICERCA\nSANITARIA",
        "display_short": "Ric. Sanitaria",
        "sheet": "Sanita",
        "slug": "Ricerca_sanitaria",
        "has_redistribuzione": True,
    },
    {
        "key": "comuni",
        "display": "\nCOMUNI",
        "display_short": "Comuni",
        "sheet": "Comuni",
        "slug": "Comuni",
        "has_redistribuzione": False,
    },
    {
        "key": "cultura",
        "display": "Enti della\nCULTURA",
        "display_short": "Cultura",
        "sheet": "Cultura",
        "slug": "Beni_culturali",
        "has_redistribuzione": False,
    },
    {
        "key": "aree_protette",
        "display": "Enti delle\nAREE\nPROTETTE",
        "display_short": "Aree Protette",
        "sheet": "Aree_Protette",
        "slug": "Aree_protette",
        "has_redistribuzione": False,
    },
]

# Categorie con redistribuzione <100 EUR
CATS_WITH_REDIST = [c["key"] for c in CATEGORIES if c["has_redistribuzione"]]

# Vecchi slug → slug canonico (per retrocompatibilita' con cartelle non rinominate)
SLUG_ALIASES = {
    "ETS_ONLUS": ["Enti_del_volontariato", "Volontariato"],
    "ASD": ["Associazioni_Sportive_Dilettantistiche_ASD"],
    "Beni_culturali": ["Beni_culturali_e_paesaggistici"],
    "Aree_protette": ["Enti_gestori_delle_aree_protette"],
}


# ============================================================================
# CARICAMENTO DATI
# ============================================================================

def find_category_files(dati_dir, anno, cat):
    """
    Cerca i file ammessi/esclusi per una categoria.
    Prova prima lo slug canonico, poi i vecchi nomi come fallback.
    Restituisce {"ammessi": path|None, "esclusi": path|None}.
    """
    result = {"ammessi": None, "esclusi": None}
    canonical = cat["slug"]
    slugs_to_try = [canonical] + SLUG_ALIASES.get(canonical, [])

    for slug in slugs_to_try:
        cat_dir = os.path.join(dati_dir, slug)
        if not os.path.isdir(cat_dir):
            continue
        for tipo in ("ammessi", "esclusi"):
            if result[tipo] is None:
                path = os.path.join(cat_dir, f"{anno}_{slug}_{tipo}.xlsx")
                if os.path.isfile(path):
                    result[tipo] = path
        # Se trovati entrambi, esci
        if result["ammessi"] and result["esclusi"]:
            break

    return result


def load_category_file(path, cat_key):
    """
    Legge un file .xlsx di categoria e restituisce una lista di dict normalizzati.
    Gestisce i 3 schemi: 12 colonne, 10 colonne, 7 colonne (Comuni).
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    # Riga 1 = header
    header = [str(c).strip() if c else "" for c in rows[0]]
    header_lower = [h.lower() for h in header]
    n_cols = len(header)

    # Rileva schema
    is_comuni = cat_key == "comuni"

    # Mappa colonne per posizione in base allo schema
    records = []
    for row in rows[1:]:
        cells = list(row) + [None] * max(0, n_cols - len(row))

        # Codice fiscale (col 1 in tutti gli schemi)
        raw_cf = str(cells[1]).strip() if cells[1] else ""
        # Rimuovi apici dal CF (formato "'01199250158'")
        raw_cf = raw_cf.strip("'\"")
        cf = normalize_cf(raw_cf)

        # Salta righe senza CF valido (disclaimer, righe vuote)
        if not cf or len(cf) < 5:
            continue
        # Salta righe che sono disclaimer
        if cf.startswith("*") or (cells[0] and str(cells[0]).startswith("*")):
            continue

        rec = {
            "cf": cf,
            "denominazione": "",
            "comune": "",
            "pr": "",
            "regione": "",
            "n_scelte": 0,
            "importo_espresse": 0.0,
            "importo_generiche": 0.0,
            "redistribuzione": 0.0,
            "totale_erogabile": 0.0,
        }

        if is_comuni:
            # Schema 7 col: Prog, CF, Comune, Provincia, Regione, Scelte, Importo
            rec["denominazione"] = str(cells[2]).strip() if cells[2] else ""
            rec["comune"] = str(cells[2]).strip() if cells[2] else ""
            rec["pr"] = str(cells[3]).strip() if cells[3] else ""
            rec["regione"] = str(cells[4]).strip() if cells[4] else ""
            rec["n_scelte"] = parse_intero(cells[5]) or 0
            importo = parse_importo(cells[6]) or 0.0
            rec["importo_espresse"] = importo
            rec["importo_generiche"] = 0.0
            rec["redistribuzione"] = 0.0
            rec["totale_erogabile"] = importo
        elif n_cols >= 12:
            # Schema 12 col: con redistribuzione ed erogabile
            rec["denominazione"] = str(cells[2]).strip() if cells[2] else ""
            rec["regione"] = str(cells[3]).strip() if cells[3] else ""
            rec["pr"] = str(cells[4]).strip() if cells[4] else ""
            rec["comune"] = str(cells[5]).strip() if cells[5] else ""
            rec["n_scelte"] = parse_intero(cells[6]) or 0
            rec["importo_espresse"] = parse_importo(cells[7]) or 0.0
            rec["importo_generiche"] = parse_importo(cells[8]) or 0.0
            rec["redistribuzione"] = parse_importo(cells[10]) or 0.0
            rec["totale_erogabile"] = parse_importo(cells[11]) or 0.0
        else:
            # Schema 10 col: senza redistribuzione
            rec["denominazione"] = str(cells[2]).strip() if cells[2] else ""
            rec["regione"] = str(cells[3]).strip() if cells[3] else ""
            rec["pr"] = str(cells[4]).strip() if cells[4] else ""
            rec["comune"] = str(cells[5]).strip() if cells[5] else ""
            rec["n_scelte"] = parse_intero(cells[6]) or 0
            rec["importo_espresse"] = parse_importo(cells[7]) or 0.0
            rec["importo_generiche"] = parse_importo(cells[8]) or 0.0
            rec["redistribuzione"] = 0.0
            totale = parse_importo(cells[9]) or 0.0
            rec["totale_erogabile"] = totale

        records.append(rec)

    return records


def load_year_data(dati_dir, anno):
    """
    Carica tutte le 7 categorie per un anno e costruisce il dict master.
    Restituisce {cf: entity_dict}.
    """
    entities = {}

    for cat in CATEGORIES:
        files = find_category_files(dati_dir, anno, cat)
        cat_key = cat["key"]

        for tipo in ("ammessi", "esclusi"):
            path = files[tipo]
            if path is None:
                continue

            status = "Ammesso" if tipo == "ammessi" else "Escluso"
            records = load_category_file(path, cat_key)
            logging.info(f"  [{anno}/{cat['display_short']}/{tipo}] {len(records)} entita'")

            for rec in records:
                cf = rec["cf"]
                if cf not in entities:
                    entities[cf] = {
                        "cf": cf,
                        "denominazione": rec["denominazione"],
                        "comune": rec["comune"],
                        "pr": rec["pr"],
                        "regione": rec["regione"],
                        "categories": {},
                        # Campi totali (calcolati dopo)
                        "totale_n_scelte": 0,
                        "totale_importo_espresse": 0.0,
                        "totale_importo_generiche": 0.0,
                        "totale_espresse_generiche": 0.0,
                        "totale_redistribuzione": 0.0,
                        "totale_erogabile": 0.0,
                        "media_scelte_espresse": 0.0,
                        "status": "",
                        "note": "",
                        # Classifica
                        "rank_importo": None,
                        "rank_scelte": None,
                        "rank_media": None,
                        # YoY
                        "yoy_importo_prev": None,
                        "yoy_importo_pct": None,
                        "yoy_scelte_prev": None,
                        "yoy_scelte_pct": None,
                        "yoy_media_prev": None,
                        "yoy_media_pct": None,
                    }
                else:
                    # Aggiorna anagrafica se mancante
                    ent = entities[cf]
                    if not ent["denominazione"] and rec["denominazione"]:
                        ent["denominazione"] = rec["denominazione"]
                    if not ent["comune"] and rec["comune"]:
                        ent["comune"] = rec["comune"]
                    if not ent["pr"] and rec["pr"]:
                        ent["pr"] = rec["pr"]
                    if not ent["regione"] and rec["regione"]:
                        ent["regione"] = rec["regione"]

                entities[cf]["categories"][cat_key] = {
                    "n_scelte": rec["n_scelte"],
                    "importo_espresse": rec["importo_espresse"],
                    "importo_generiche": rec["importo_generiche"],
                    "redistribuzione": rec["redistribuzione"],
                    "totale_erogabile": rec["totale_erogabile"],
                    "status": status,
                }

    logging.info(f"  Totale entita' uniche: {len(entities)}")
    return entities


# ============================================================================
# CALCOLI DERIVATI
# ============================================================================

def compute_totals(entities):
    """Calcola totali, media e status globale per ogni entita'."""
    for ent in entities.values():
        tot_scelte = 0
        tot_espresse = 0.0
        tot_generiche = 0.0
        tot_redist = 0.0
        tot_erogabile = 0.0
        has_ammesso = False

        for cat_key, cat_data in ent["categories"].items():
            tot_scelte += cat_data["n_scelte"]
            tot_espresse += cat_data["importo_espresse"]
            tot_generiche += cat_data["importo_generiche"]
            tot_redist += cat_data["redistribuzione"]
            tot_erogabile += cat_data["totale_erogabile"]
            if cat_data["status"] == "Ammesso":
                has_ammesso = True

        ent["totale_n_scelte"] = tot_scelte
        ent["totale_importo_espresse"] = tot_espresse
        ent["totale_importo_generiche"] = tot_generiche
        ent["totale_espresse_generiche"] = tot_espresse + tot_generiche
        ent["totale_redistribuzione"] = tot_redist
        ent["totale_erogabile"] = tot_erogabile
        ent["media_scelte_espresse"] = (
            tot_espresse / tot_scelte if tot_scelte > 0 else 0.0
        )
        ent["status"] = "Ammesso" if has_ammesso else "Escluso"


def compute_rankings(entities):
    """Assegna classifica per importo, scelte e media (solo ammessi)."""
    ammessi = [e for e in entities.values() if e["status"] == "Ammesso"]

    # Per importo erogabile
    by_importo = sorted(ammessi, key=lambda e: e["totale_erogabile"], reverse=True)
    for i, e in enumerate(by_importo, 1):
        e["rank_importo"] = i

    # Per numero scelte
    by_scelte = sorted(ammessi, key=lambda e: e["totale_n_scelte"], reverse=True)
    for i, e in enumerate(by_scelte, 1):
        e["rank_scelte"] = i

    # Per media scelte espresse
    by_media = sorted(ammessi, key=lambda e: e["media_scelte_espresse"], reverse=True)
    for i, e in enumerate(by_media, 1):
        e["rank_media"] = i


def compute_yoy(current, previous):
    """Calcola confronto anno su anno per ogni entita' corrente."""
    for cf, ent in current.items():
        if cf in previous:
            prev = previous[cf]
            # Importo
            ent["yoy_importo_prev"] = prev["totale_erogabile"]
            if prev["totale_erogabile"] > 0:
                ent["yoy_importo_pct"] = (
                    (ent["totale_erogabile"] - prev["totale_erogabile"])
                    / prev["totale_erogabile"]
                )
            # Scelte
            ent["yoy_scelte_prev"] = prev["totale_n_scelte"]
            if prev["totale_n_scelte"] > 0:
                ent["yoy_scelte_pct"] = (
                    (ent["totale_n_scelte"] - prev["totale_n_scelte"])
                    / prev["totale_n_scelte"]
                )
            # Media
            ent["yoy_media_prev"] = prev["media_scelte_espresse"]
            if prev["media_scelte_espresse"] > 0:
                ent["yoy_media_pct"] = (
                    (ent["media_scelte_espresse"] - prev["media_scelte_espresse"])
                    / prev["media_scelte_espresse"]
                )


def compute_regional_stats(entities):
    """
    Aggrega per regione: nr. enti ammessi/esclusi, nr. scelte, importo erogato
    per ogni categoria.
    """
    regions = {}
    for ent in entities.values():
        regione = ent["regione"] or "SCONOSCIUTA"
        if regione not in regions:
            regions[regione] = {}
            for cat in CATEGORIES:
                regions[regione][cat["key"]] = {
                    "n_ammessi": 0, "n_esclusi": 0,
                    "n_scelte": 0, "importo": 0.0,
                }
        for cat_key, cat_data in ent["categories"].items():
            r = regions[regione][cat_key]
            if cat_data["status"] == "Ammesso":
                r["n_ammessi"] += 1
            else:
                r["n_esclusi"] += 1
            r["n_scelte"] += cat_data["n_scelte"]
            r["importo"] += cat_data["totale_erogabile"]

    # Ordina per nome regione
    return dict(sorted(regions.items()))


# ============================================================================
# FORMATTAZIONE EXCEL
# ============================================================================

def create_formats(wb):
    """Crea tutti i formati riutilizzabili per xlsxwriter."""
    ec = EXCEL_CONFIG
    base_header = {
        "font_name": ec["header_font"],
        "bold": True,
        "font_size": ec["header_size"],
        "font_color": ec["header_fg"],
        "bg_color": ec["header_bg"],
        "align": "center",
        "valign": "vcenter",
        "text_wrap": True,
        "border": 1,
        "border_color": ec["border_color"],
    }
    base_data = {
        "font_name": ec["data_font"],
        "font_size": ec["data_size"],
        "border": 1,
        "border_color": ec["border_color"],
        "valign": "vcenter",
    }
    return {
        "header": wb.add_format(base_header),
        "header_group": wb.add_format({
            **base_header, "font_size": ec["header_size"] + 1, "bottom": 0,
        }),
        "title": wb.add_format({
            "font_name": ec["header_font"], "bold": True, "font_size": 14,
            "valign": "vcenter",
        }),
        "subtitle": wb.add_format({
            "font_name": ec["data_font"], "font_size": 10, "italic": True,
            "valign": "vcenter",
        }),
        "summary": wb.add_format({
            **base_data, "bold": True, "bg_color": "#D9E2F3",
            "num_format": "#,##0.00",
        }),
        "summary_int": wb.add_format({
            **base_data, "bold": True, "bg_color": "#D9E2F3",
            "num_format": "#,##0",
        }),
        "data": wb.add_format(base_data),
        "data_alt": wb.add_format({**base_data, "bg_color": ec["alt_row_bg"]}),
        "data_int": wb.add_format({**base_data, "num_format": "#,##0"}),
        "data_int_alt": wb.add_format({
            **base_data, "bg_color": ec["alt_row_bg"], "num_format": "#,##0",
        }),
        "data_euro": wb.add_format({**base_data, "num_format": "#,##0.00"}),
        "data_euro_alt": wb.add_format({
            **base_data, "bg_color": ec["alt_row_bg"], "num_format": "#,##0.00",
        }),
        "data_pct": wb.add_format({**base_data, "num_format": "0.00%"}),
        "data_pct_alt": wb.add_format({
            **base_data, "bg_color": ec["alt_row_bg"], "num_format": "0.00%",
        }),
        "ammesso": wb.add_format({**base_data, "font_color": "#006100"}),
        "ammesso_alt": wb.add_format({
            **base_data, "bg_color": ec["alt_row_bg"], "font_color": "#006100",
        }),
        "escluso": wb.add_format({**base_data, "font_color": "#9C0006"}),
        "escluso_alt": wb.add_format({
            **base_data, "bg_color": ec["alt_row_bg"], "font_color": "#9C0006",
        }),
    }


# ============================================================================
# SCRITTURA FOGLIO COMPLESSIVO
# ============================================================================

def _build_complessivo_columns(has_yoy):
    """Costruisce la lista delle colonne per il foglio COMPLESSIVO."""
    cols = []

    # Anagrafica (0-4)
    cols.append({"name": "CODICE FISCALE", "group": "", "width": 16, "type": "text"})
    cols.append({"name": "DENOMINAZIONE", "group": "", "width": 40, "type": "text"})
    cols.append({"name": "COMUNE", "group": "", "width": 18, "type": "text"})
    cols.append({"name": "PR", "group": "", "width": 5, "type": "text"})
    cols.append({"name": "REGIONE", "group": "", "width": 16, "type": "text"})

    # Nr. Scelte per categoria (5-11)
    for cat in CATEGORIES:
        cols.append({
            "name": cat["display"], "group": "NUMERO SCELTE ESPRESSE",
            "width": 12, "type": "int", "cat_key": cat["key"], "field": "n_scelte",
        })

    # Importo Espresse per categoria (12-18)
    for cat in CATEGORIES:
        cols.append({
            "name": cat["display"], "group": "IMPORTO SCELTE ESPRESSE",
            "width": 14, "type": "euro", "cat_key": cat["key"], "field": "importo_espresse",
        })

    # Importo Generiche per categoria (19-25)
    for cat in CATEGORIES:
        cols.append({
            "name": cat["display"], "group": "IMPORTO SCELTE GENERICHE",
            "width": 14, "type": "euro", "cat_key": cat["key"], "field": "importo_generiche",
        })

    # Redistribuzione <100 EUR (26-29) — solo 4 categorie
    for cat in CATEGORIES:
        if cat["has_redistribuzione"]:
            cols.append({
                "name": cat["display"], "group": "REDISTRIBUZIONE\nIMPORTI <100 EUR",
                "width": 14, "type": "euro", "cat_key": cat["key"], "field": "redistribuzione",
            })

    # Totali (30-38)
    cols.append({"name": "TOTALE\nSCELTE\nESPRESSE", "group": "TOTALE DESTINAZIONI",
                 "width": 12, "type": "int", "field": "totale_n_scelte"})
    cols.append({"name": "TOTALE\nSCELTE\nESPRESSE", "group": "TOTALE DESTINAZIONI",
                 "width": 14, "type": "euro", "field": "totale_importo_espresse"})
    cols.append({"name": "TOTALE\nSCELTE\nGENERICHE", "group": "TOTALE DESTINAZIONI",
                 "width": 14, "type": "euro", "field": "totale_importo_generiche"})
    cols.append({"name": "TOTALE\n(Espresse+Generiche)", "group": "TOTALE DESTINAZIONI",
                 "width": 14, "type": "euro", "field": "totale_espresse_generiche"})
    cols.append({"name": "TOTALE\nREDISTRIBUZIONE", "group": "TOTALE DESTINAZIONI",
                 "width": 14, "type": "euro", "field": "totale_redistribuzione"})
    cols.append({"name": "TOTALE\nEROGABILE", "group": "TOTALE DESTINAZIONI",
                 "width": 14, "type": "euro", "field": "totale_erogabile"})
    cols.append({"name": "MEDIA\nSCELTE ESPRESSE", "group": "TOTALE DESTINAZIONI",
                 "width": 12, "type": "euro", "field": "media_scelte_espresse"})

    # Status
    cols.append({"name": "Ammessi ed\nEsclusi", "group": "", "width": 10, "type": "status", "field": "status"})
    cols.append({"name": "NOTE", "group": "", "width": 12, "type": "text", "field": "note"})

    # Classifica
    cols.append({"name": "Posizione x\nImporto Totale", "group": "CLASSIFICA",
                 "width": 10, "type": "int", "field": "rank_importo"})
    cols.append({"name": "Posizione x\nNr. Scelte", "group": "CLASSIFICA",
                 "width": 10, "type": "int", "field": "rank_scelte"})
    cols.append({"name": "Posizione x\nValore Medio", "group": "CLASSIFICA",
                 "width": 10, "type": "int", "field": "rank_media"})

    # Confronto YoY
    if has_yoy:
        cols.append({"name": "Importo Totale\nAnno Prec.", "group": "CONFRONTO ANNO PRECEDENTE",
                     "width": 14, "type": "euro", "field": "yoy_importo_prev"})
        cols.append({"name": "Importo\nVariazione %", "group": "CONFRONTO ANNO PRECEDENTE",
                     "width": 10, "type": "pct", "field": "yoy_importo_pct"})
        cols.append({"name": "Nr. Scelte\nAnno Prec.", "group": "CONFRONTO ANNO PRECEDENTE",
                     "width": 12, "type": "int", "field": "yoy_scelte_prev"})
        cols.append({"name": "Nr. Scelte\nVariazione %", "group": "CONFRONTO ANNO PRECEDENTE",
                     "width": 10, "type": "pct", "field": "yoy_scelte_pct"})
        cols.append({"name": "Valore Medio\nAnno Prec.", "group": "CONFRONTO ANNO PRECEDENTE",
                     "width": 12, "type": "euro", "field": "yoy_media_prev"})
        cols.append({"name": "Valore Medio\nVariazione %", "group": "CONFRONTO ANNO PRECEDENTE",
                     "width": 10, "type": "pct", "field": "yoy_media_pct"})

    return cols


def _get_cell_value(ent, col):
    """Estrai il valore di una cella per un'entita' in base alla definizione colonna."""
    field = col.get("field")
    cat_key = col.get("cat_key")

    if field is None:
        return ""

    if cat_key:
        # Campo per-categoria
        cat_data = ent["categories"].get(cat_key)
        if cat_data is None:
            return None
        return cat_data.get(field, 0)
    else:
        return ent.get(field)


def _write_cell(ws, row, col_idx, value, col_def, fmts, is_alt):
    """Scrive una cella con il formato appropriato."""
    ctype = col_def["type"]

    if value is None:
        ws.write_blank(row, col_idx, "", fmts["data_alt"] if is_alt else fmts["data"])
        return

    if ctype == "text":
        fmt = fmts["data_alt"] if is_alt else fmts["data"]
        ws.write_string(row, col_idx, str(value) if value else "", fmt)
    elif ctype == "int":
        fmt = fmts["data_int_alt"] if is_alt else fmts["data_int"]
        if isinstance(value, (int, float)) and value:
            ws.write_number(row, col_idx, int(value), fmt)
        else:
            ws.write_blank(row, col_idx, "", fmt)
    elif ctype == "euro":
        fmt = fmts["data_euro_alt"] if is_alt else fmts["data_euro"]
        if isinstance(value, (int, float)) and value:
            ws.write_number(row, col_idx, float(value), fmt)
        else:
            ws.write_blank(row, col_idx, "", fmt)
    elif ctype == "pct":
        fmt = fmts["data_pct_alt"] if is_alt else fmts["data_pct"]
        if isinstance(value, (int, float)) and value is not None:
            ws.write_number(row, col_idx, float(value), fmt)
        else:
            ws.write_blank(row, col_idx, "", fmt)
    elif ctype == "status":
        txt = str(value) if value else ""
        if txt == "Ammesso":
            fmt = fmts["ammesso_alt"] if is_alt else fmts["ammesso"]
        elif txt == "Escluso":
            fmt = fmts["escluso_alt"] if is_alt else fmts["escluso"]
        else:
            fmt = fmts["data_alt"] if is_alt else fmts["data"]
        ws.write_string(row, col_idx, txt, fmt)
    else:
        ws.write(row, col_idx, value, fmts["data_alt"] if is_alt else fmts["data"])


def write_complessivo(wb, entities, anno, fmts, has_yoy, anno_confronto=None):
    """Scrive il foglio COMPLESSIVO."""
    ws = wb.add_worksheet("COMPLESSIVO")
    cols = _build_complessivo_columns(has_yoy)

    # Ordina entita': ammessi prima (per importo desc), poi esclusi
    sorted_ents = sorted(
        entities.values(),
        key=lambda e: (0 if e["status"] == "Ammesso" else 1, -e["totale_erogabile"]),
    )

    # --- Righe di titolo (0-3) ---
    ws.write(0, 1, f"Elenco iscritti al riparto del 5xMille", fmts["title"])
    anno_imp = anno - 1
    ws.write(1, 1, f"Anno Finanziario {anno} (Anno d'Imposta {anno_imp})", fmts["subtitle"])
    ws.write(2, 1, f"Elaborazione automatica su dati Agenzia delle Entrate", fmts["subtitle"])

    # Sub-totali nella riga 1-3 (colonna D+)
    n_ammessi = sum(1 for e in sorted_ents if e["status"] == "Ammesso")
    n_esclusi = sum(1 for e in sorted_ents if e["status"] == "Escluso")
    ws.write(1, 4, f"TOTALE ISCRITTI: {len(sorted_ents):,}", fmts["summary"])
    ws.write(2, 4, f"TOTALE ESCLUSI: {n_esclusi:,}", fmts["summary"])
    ws.write(3, 4, f"TOTALE AMMESSI: {n_ammessi:,}", fmts["summary"])

    # --- Riga header gruppo (4) e header colonne (5) ---
    header_row = 4
    col_header_row = 5

    # Scrivi merged header gruppi
    prev_group = None
    group_start = None
    for ci, col in enumerate(cols):
        grp = col["group"]
        if grp and grp != prev_group:
            if prev_group:
                # Chiudi gruppo precedente
                if group_start < ci - 1:
                    ws.merge_range(
                        header_row, group_start, header_row, ci - 1,
                        prev_group, fmts["header_group"],
                    )
                else:
                    ws.write(header_row, group_start, prev_group, fmts["header_group"])
            group_start = ci
        elif not grp and prev_group:
            # Chiudi gruppo precedente
            if group_start < ci - 1:
                ws.merge_range(
                    header_row, group_start, header_row, ci - 1,
                    prev_group, fmts["header_group"],
                )
            else:
                ws.write(header_row, group_start, prev_group, fmts["header_group"])
            prev_group = None
            group_start = None
        prev_group = grp if grp else None
        if grp and group_start is None:
            group_start = ci

    # Chiudi ultimo gruppo
    if prev_group and group_start is not None:
        last = len(cols) - 1
        if group_start < last:
            ws.merge_range(
                header_row, group_start, header_row, last,
                prev_group, fmts["header_group"],
            )
        else:
            ws.write(header_row, group_start, prev_group, fmts["header_group"])

    # Scrivi header colonne
    for ci, col in enumerate(cols):
        ws.write(col_header_row, ci, col["name"], fmts["header"])
        ws.set_column(ci, ci, col["width"])

    # --- Dati (riga 6+) ---
    data_start = 6
    for ri, ent in enumerate(sorted_ents):
        row = data_start + ri
        is_alt = ri % 2 == 1

        for ci, col in enumerate(cols):
            if col.get("field") is None and col.get("cat_key") is None:
                # Colonna anagrafica senza field — usa nome colonna per lookup
                name = col["name"]
                if name == "CODICE FISCALE":
                    val = ent["cf"]
                elif name == "DENOMINAZIONE":
                    val = ent["denominazione"]
                elif name == "COMUNE":
                    val = ent["comune"]
                elif name == "PR":
                    val = ent["pr"]
                elif name == "REGIONE":
                    val = ent["regione"]
                else:
                    val = ""
            else:
                val = _get_cell_value(ent, col)

            _write_cell(ws, row, ci, val, col, fmts, is_alt)

    # Freeze e autofilter
    ws.freeze_panes(data_start, 5)
    ws.autofilter(col_header_row, 0, col_header_row + len(sorted_ents), len(cols) - 1)

    logging.info(f"  COMPLESSIVO: {len(sorted_ents)} righe, {len(cols)} colonne")


# ============================================================================
# SCRITTURA FOGLI PER CATEGORIA
# ============================================================================

def write_category_sheet(wb, entities, cat, anno, fmts, has_yoy):
    """Scrive un foglio per una singola categoria."""
    cat_key = cat["key"]
    sheet_name = cat["sheet"][:31]

    # Filtra entita' che hanno dati per questa categoria
    cat_ents = []
    for ent in entities.values():
        if cat_key in ent["categories"]:
            cat_ents.append(ent)

    if not cat_ents:
        logging.info(f"  {sheet_name}: nessuna entita', foglio saltato")
        return

    # Ordina: ammessi (per erogabile desc) poi esclusi
    cat_ents.sort(key=lambda e: (
        0 if e["categories"][cat_key]["status"] == "Ammesso" else 1,
        -e["categories"][cat_key]["totale_erogabile"],
    ))

    # Calcola classifica interna alla categoria
    ammessi = [e for e in cat_ents if e["categories"][cat_key]["status"] == "Ammesso"]
    by_importo = sorted(ammessi, key=lambda e: e["categories"][cat_key]["totale_erogabile"], reverse=True)
    by_scelte = sorted(ammessi, key=lambda e: e["categories"][cat_key]["n_scelte"], reverse=True)
    by_media = sorted(ammessi, key=lambda e: (
        e["categories"][cat_key]["importo_espresse"] / e["categories"][cat_key]["n_scelte"]
        if e["categories"][cat_key]["n_scelte"] > 0 else 0
    ), reverse=True)

    cat_rank_importo = {e["cf"]: i + 1 for i, e in enumerate(by_importo)}
    cat_rank_scelte = {e["cf"]: i + 1 for i, e in enumerate(by_scelte)}
    cat_rank_media = {e["cf"]: i + 1 for i, e in enumerate(by_media)}

    ws = wb.add_worksheet(sheet_name)

    # Definisci colonne
    col_defs = [
        {"name": "CODICE FISCALE", "width": 16, "type": "text"},
        {"name": "DENOMINAZIONE", "width": 40, "type": "text"},
        {"name": "COMUNE", "width": 18, "type": "text"},
        {"name": "PR", "width": 5, "type": "text"},
        {"name": "REGIONE", "width": 16, "type": "text"},
        {"name": "NUMERO\nSCELTE ESPRESSE", "width": 12, "type": "int"},
        {"name": "IMPORTO\nSCELTE ESPRESSE", "width": 14, "type": "euro"},
        {"name": "IMPORTO\nSCELTE GENERICHE", "width": 14, "type": "euro"},
        {"name": "IMPORTO TOTALE\n(Espresse+Generiche)", "width": 14, "type": "euro"},
    ]
    if cat["has_redistribuzione"]:
        col_defs.append({"name": "REDISTRIBUZIONE\n<100 EUR", "width": 14, "type": "euro"})
    col_defs.extend([
        {"name": "TOTALE\nEROGABILE", "width": 14, "type": "euro"},
        {"name": "MEDIA\nSCELTE ESPRESSE", "width": 12, "type": "euro"},
        {"name": "Ammessi ed\nEsclusi", "width": 10, "type": "status"},
        {"name": "NOTE", "width": 12, "type": "text"},
        {"name": "Pos. x\nImporto", "width": 9, "type": "int"},
        {"name": "Pos. x\nScelte", "width": 9, "type": "int"},
        {"name": "Pos. x\nMedia", "width": 9, "type": "int"},
    ])
    if has_yoy:
        col_defs.extend([
            {"name": "Importo Totale\nAnno Prec.", "width": 14, "type": "euro"},
            {"name": "Importo\nVar. %", "width": 10, "type": "pct"},
            {"name": "Nr. Scelte\nAnno Prec.", "width": 12, "type": "int"},
            {"name": "Nr. Scelte\nVar. %", "width": 10, "type": "pct"},
            {"name": "Media\nAnno Prec.", "width": 12, "type": "euro"},
            {"name": "Media\nVar. %", "width": 10, "type": "pct"},
        ])

    # Righe di riepilogo
    n_amm = sum(1 for e in cat_ents if e["categories"][cat_key]["status"] == "Ammesso")
    n_esc = sum(1 for e in cat_ents if e["categories"][cat_key]["status"] == "Escluso")
    ws.write(0, 0, f"{cat['display_short']} - Anno Finanziario {anno}", fmts["title"])
    ws.write(1, 0, f"Ammessi: {n_amm:,}   Esclusi: {n_esc:,}   Totale: {len(cat_ents):,}", fmts["subtitle"])

    # Header
    header_row = 3
    for ci, col in enumerate(col_defs):
        ws.write(header_row, ci, col["name"], fmts["header"])
        ws.set_column(ci, ci, col["width"])

    # Dati
    data_start = 4
    for ri, ent in enumerate(cat_ents):
        row = data_start + ri
        is_alt = ri % 2 == 1
        cd = ent["categories"][cat_key]
        media = cd["importo_espresse"] / cd["n_scelte"] if cd["n_scelte"] > 0 else 0.0
        espresse_gen = cd["importo_espresse"] + cd["importo_generiche"]

        vals = [
            ent["cf"], ent["denominazione"], ent["comune"], ent["pr"], ent["regione"],
            cd["n_scelte"], cd["importo_espresse"], cd["importo_generiche"],
            espresse_gen,
        ]
        if cat["has_redistribuzione"]:
            vals.append(cd["redistribuzione"])
        vals.extend([
            cd["totale_erogabile"], media, cd["status"], ent["note"],
            cat_rank_importo.get(ent["cf"]),
            cat_rank_scelte.get(ent["cf"]),
            cat_rank_media.get(ent["cf"]),
        ])
        if has_yoy:
            vals.extend([
                ent.get("yoy_importo_prev"),
                ent.get("yoy_importo_pct"),
                ent.get("yoy_scelte_prev"),
                ent.get("yoy_scelte_pct"),
                ent.get("yoy_media_prev"),
                ent.get("yoy_media_pct"),
            ])

        for ci, val in enumerate(vals):
            _write_cell(ws, row, ci, val, col_defs[ci], fmts, is_alt)

    # Freeze e autofilter
    ws.freeze_panes(data_start, 5)
    ws.autofilter(header_row, 0, header_row + len(cat_ents), len(col_defs) - 1)

    logging.info(f"  {sheet_name}: {len(cat_ents)} righe")


# ============================================================================
# SCRITTURA FOGLIO GRAFICI
# ============================================================================

def write_grafici(wb, regional, anno, fmts):
    """Scrive il foglio con l'aggregazione regionale."""
    ws = wb.add_worksheet("Grafici")

    # Header gruppo
    col_defs = [{"name": "Regione", "width": 18, "type": "text"}]

    # Per ogni categoria: Nr. Enti Ammessi, Nr. Enti Esclusi, Nr. Scelte, Importo
    for cat in CATEGORIES:
        prefix = cat["display_short"]
        col_defs.extend([
            {"name": f"{prefix}\nNr. Ammessi", "width": 12, "type": "int", "cat_key": cat["key"], "sub": "n_ammessi"},
            {"name": f"{prefix}\nNr. Esclusi", "width": 12, "type": "int", "cat_key": cat["key"], "sub": "n_esclusi"},
            {"name": f"{prefix}\nNr. Scelte", "width": 12, "type": "int", "cat_key": cat["key"], "sub": "n_scelte"},
            {"name": f"{prefix}\nImporto", "width": 14, "type": "euro", "cat_key": cat["key"], "sub": "importo"},
        ])

    # Header
    header_row = 0
    for ci, col in enumerate(col_defs):
        ws.write(header_row, ci, col["name"], fmts["header"])
        ws.set_column(ci, ci, col["width"])

    # Dati
    data_start = 1
    for ri, (regione, reg_data) in enumerate(regional.items()):
        row = data_start + ri
        is_alt = ri % 2 == 1
        fmt_t = fmts["data_alt"] if is_alt else fmts["data"]
        ws.write_string(row, 0, regione, fmt_t)

        for ci, col in enumerate(col_defs[1:], 1):
            cat_stats = reg_data.get(col["cat_key"], {})
            val = cat_stats.get(col["sub"], 0)
            _write_cell(ws, row, ci, val if val else None, col, fmts, is_alt)

    # Freeze
    ws.freeze_panes(data_start, 1)
    ws.autofilter(header_row, 0, header_row + len(regional), len(col_defs) - 1)
    logging.info(f"  Grafici: {len(regional)} regioni")


# ============================================================================
# SCRITTURA FOGLIO LICENZA
# ============================================================================

def write_licenza(wb, anno, fmts):
    """Scrive il foglio di attribuzione."""
    ws = wb.add_worksheet("Licenza")
    ws.set_column(0, 0, 80)

    lines = [
        f"Report 5 per Mille - Anno Finanziario {anno}",
        "",
        "Questa elaborazione e' basata sui dati ufficiali pubblicati dall'Agenzia delle Entrate",
        f"relativi al riparto del 5 per mille dell'anno finanziario {anno}.",
        "",
        "Fonte dati: Agenzia delle Entrate - www.agenziaentrate.gov.it",
        "",
        "Struttura report ispirata al lavoro di Nicola Bedogni per ASSIF.",
        "",
        "Report generato automaticamente da report.py",
        "Progetto: github.com/aalmagio/5x1000",
    ]

    for i, line in enumerate(lines):
        if i == 0:
            ws.write(i, 0, line, fmts["title"])
        else:
            ws.write(i, 0, line, fmts["subtitle"])

    logging.info(f"  Licenza: OK")


# ============================================================================
# CLI
# ============================================================================

def parse_args():
    parser = argparse.ArgumentParser(
        description="5 per Mille - Genera report Excel (stile ASSIF/Bedogni)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""\
Esempi:
  %(prog)s --anno 2024,2023                    # report 2024, confronto con 2023
  %(prog)s --anno 2024                         # report 2024, confronto con 2023 (anno-1)
  %(prog)s --anno 2024 --no-confronto          # report 2024 senza confronto YoY
  %(prog)s --anno 2024 --output report.xlsx    # output personalizzato
""",
    )
    parser.add_argument(
        "--anno", type=str, required=True,
        help="Anno/i per il report: '2024' (confronto con anno-1) o '2024,2023' (confronto esplicito)",
    )
    parser.add_argument(
        "--anno-confronto", type=int, default=None,
        help="Anno per il confronto YoY (alternativa a '2024,2023')",
    )
    parser.add_argument(
        "--no-confronto", action="store_true",
        help="Non calcolare il confronto con l'anno precedente",
    )
    parser.add_argument(
        "--output", type=str, default=None,
        help="Percorso file output (default: Dati/report_{anno}.xlsx)",
    )
    parser.add_argument(
        "--root", type=str, default=".",
        help="Cartella root del progetto",
    )

    args = parser.parse_args()

    # Parsing --anno: "2024" o "2024,2023"
    parts = [p.strip() for p in args.anno.split(",") if p.strip()]
    try:
        args.anno = int(parts[0])
    except ValueError:
        parser.error(f"Anno non valido: '{parts[0]}'")
    if len(parts) >= 2:
        try:
            args.anno_confronto = int(parts[1])
        except ValueError:
            parser.error(f"Anno confronto non valido: '{parts[1]}'")

    return args


# ============================================================================
# MAIN
# ============================================================================

def main():
    args = parse_args()

    root_dir = Path(args.root).resolve()

    # Config e logging
    cfg = _load_config(root_dir)
    _apply_config(cfg)
    log_file = setup_logging(root_dir)
    logging.info(f"=== Report 5xMille - Anno {args.anno} ===")
    logging.info(f"Directory root: {root_dir}")

    dati_dir = str(root_dir / "Dati")
    anno = args.anno

    # 1. Carica anno corrente
    logging.info(f"Caricamento dati anno {anno}...")
    entities = load_year_data(dati_dir, anno)
    if not entities:
        logging.error(f"Nessun dato trovato per l'anno {anno}")
        sys.exit(1)

    # 2. Calcola totali e classifiche
    compute_totals(entities)
    compute_rankings(entities)

    # 3. Confronto anno precedente
    has_yoy = False
    anno_confronto = None
    if not args.no_confronto:
        anno_confronto = args.anno_confronto or (anno - 1)
        logging.info(f"Caricamento dati anno confronto {anno_confronto}...")
        entities_prev = load_year_data(dati_dir, anno_confronto)
        if entities_prev:
            compute_totals(entities_prev)
            compute_yoy(entities, entities_prev)
            has_yoy = True
            logging.info(f"  Confronto YoY: {sum(1 for e in entities.values() if e['yoy_importo_prev'] is not None)} match")
        else:
            logging.warning(f"Nessun dato per anno confronto {anno_confronto}, YoY disabilitato")

    # 4. Aggregazione regionale
    regional = compute_regional_stats(entities)

    # 5. Scrivi Excel
    output_path = args.output or os.path.join(dati_dir, f"report_{anno}.xlsx")
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    logging.info(f"Scrittura report: {output_path}")

    wb = xlsxwriter.Workbook(output_path, {"constant_memory": False})
    fmts = create_formats(wb)

    # COMPLESSIVO
    write_complessivo(wb, entities, anno, fmts, has_yoy, anno_confronto)

    # Fogli per categoria
    for cat in CATEGORIES:
        write_category_sheet(wb, entities, cat, anno, fmts, has_yoy)

    # Grafici
    write_grafici(wb, regional, anno, fmts)

    # Licenza
    write_licenza(wb, anno, fmts)

    wb.close()

    # Riepilogo
    file_size = os.path.getsize(output_path) / (1024 * 1024)
    logging.info(f"Report generato: {output_path} ({file_size:.1f} MB)")
    logging.info(f"Log completo: {log_file}")


if __name__ == "__main__":
    main()
