#!/usr/bin/env python3
"""
analizza_tracciati.py — Evoluzione dei tracciati PDF/CSV 5x1000

Produce due report distinti (o uno solo, a scelta):

  REPORT GENERALE  (dati_XXXX.xlsx)
    - Foglio "Elenco_Generale": matrice campo × anno
    - Foglio "Variazioni":      riepilogo di ogni cambiamento

  REPORT CATEGORIE  (ANNO_SLUG_ammessi/esclusi.xlsx)
    - Un foglio per ogni categoria × stato (ammessi/esclusi)
    - Foglio "Variazioni_Cat":       cambiamenti nei file di categoria
    - Foglio "Confronto_Ammessi":    tutte le categorie affiancate per anno
    - Foglio "Confronto_Esclusi":    idem per esclusi

Uso:
    python analizza_tracciati.py                        # entrambi i report, ultimi 5 anni
    python analizza_tracciati.py --tipo generale
    python analizza_tracciati.py --tipo categorie
    python analizza_tracciati.py --tipo categorie --categoria ETS_ONLUS
    python analizza_tracciati.py --anni 3 --output-dir Dati/
"""

import argparse
import difflib
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Richiesto openpyxl: pip install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Costanti
# ---------------------------------------------------------------------------

CATEGORIE_SLUG = [
    "ETS_ONLUS",
    "ASD",
    "Ricerca_scientifica",
    "Ricerca_sanitaria",
    "Comuni",
    "Beni_culturali",
    "Aree_protette",
]

C_HEADER  = "2F5496"
C_PRESENT = "D6E4BC"
C_ABSENT  = "F2F2F2"
C_NEW     = "C6EFCE"
C_REMOVED = "FFC7CE"
C_RENAMED = "FFEB9C"

# ---------------------------------------------------------------------------
# Lettura intestazioni
# ---------------------------------------------------------------------------

def get_headers(xlsx_path: Path) -> list[str]:
    try:
        import warnings
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            headers = [str(c).strip() if c is not None else "" for c in row]
            headers = [h for h in headers if h]
            if headers:
                wb.close()
                return headers
        wb.close()
    except Exception as e:
        print(f"  [WARN] {xlsx_path.name}: {e}")
    return []

# ---------------------------------------------------------------------------
# Analisi
# ---------------------------------------------------------------------------

def similarity(a: str, b: str) -> float:
    return difflib.SequenceMatcher(None, a.lower(), b.lower()).ratio()


def find_renames(removed: list[str], added: list[str], threshold: float = 0.72) -> dict[str, str]:
    renames: dict[str, str] = {}
    used: set[str] = set()
    for rem in removed:
        best_score, best_match = 0.0, None
        for add in added:
            if add in used:
                continue
            s = similarity(rem, add)
            if s > best_score:
                best_score, best_match = s, add
        if best_match and best_score >= threshold:
            renames[rem] = best_match
            used.add(best_match)
    return renames


def analizza_dataset(nome: str, headers_per_anno: dict[int, list[str]], anni: list[int]) -> dict:
    anni_presenti = [a for a in anni if a in headers_per_anno]
    all_fields: list[str] = []
    seen: set[str] = set()
    for anno in sorted(anni_presenti):
        for h in headers_per_anno[anno]:
            if h not in seen:
                all_fields.append(h)
                seen.add(h)
    matrix: dict[str, dict[int, bool]] = {
        field: {anno: field in headers_per_anno.get(anno, []) for anno in anni}
        for field in all_fields
    }
    variazioni = []
    for i in range(1, len(anni_presenti)):
        ap, ac = sorted(anni_presenti)[i - 1], sorted(anni_presenti)[i]
        hp = set(headers_per_anno.get(ap, []))
        hc = set(headers_per_anno.get(ac, []))
        added   = [h for h in headers_per_anno.get(ac, []) if h not in hp]
        removed = [h for h in headers_per_anno.get(ap, []) if h not in hc]
        renames = find_renames(removed, added)
        variazioni.append({
            "da": ap, "a": ac,
            "aggiunti": added, "rimossi": removed, "rinominati": renames,
            "n_da": len(hp), "n_a": len(hc),
        })
    return {
        "nome": nome, "anni": anni, "anni_presenti": anni_presenti,
        "all_fields": all_fields, "matrix": matrix, "variazioni": variazioni,
        "headers_per_anno": headers_per_anno,
    }

# ---------------------------------------------------------------------------
# Scrittura fogli
# ---------------------------------------------------------------------------

def _hdr_cell(ws, row, col, value):
    c = ws.cell(row, col, value)
    c.fill = PatternFill("solid", fgColor=C_HEADER)
    c.font = Font(color="FFFFFF", bold=True, size=9)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return c


def write_matrix_sheet(wb, dataset: dict, anni: list[int]):
    title = dataset["nome"][:31]
    ws = wb.create_sheet(title=title)

    f_present = PatternFill("solid", fgColor=C_PRESENT)
    f_absent  = PatternFill("solid", fgColor=C_ABSENT)
    f_new     = PatternFill("solid", fgColor=C_NEW)
    f_removed = PatternFill("solid", fgColor=C_REMOVED)
    f_renamed = PatternFill("solid", fgColor=C_RENAMED)
    align_c   = Alignment(horizontal="center", vertical="center")

    rename_old: dict[str, int] = {}
    rename_new: dict[str, int] = {}
    for v in dataset["variazioni"]:
        for old, new in v["rinominati"].items():
            rename_old[old] = v["a"]
            rename_new[new] = v["a"]

    _hdr_cell(ws, 1, 1, "Campo")
    ws.column_dimensions["A"].width = 42
    for j, anno in enumerate(anni, 2):
        _hdr_cell(ws, 1, j, str(anno))
        ws.column_dimensions[get_column_letter(j)].width = 9

    for i, field in enumerate(dataset["all_fields"], 2):
        ws.cell(i, 1, field).font = Font(size=9)
        for j, anno in enumerate(anni, 2):
            present   = dataset["matrix"][field].get(anno, False)
            prev_idx  = j - 3
            prev_anno = anni[prev_idx] if prev_idx >= 0 else None
            prev_present = dataset["matrix"][field].get(prev_anno) if prev_anno is not None else None

            c = ws.cell(i, j)
            c.alignment = align_c
            c.font = Font(size=9)

            if present:
                if field in rename_new and rename_new[field] == anno:
                    c.value, c.fill = "R↑", f_renamed
                elif prev_present is False:
                    c.value, c.fill = "+", f_new
                else:
                    c.value, c.fill = "✓", f_present
            else:
                if field in rename_old and rename_old[field] == anno:
                    c.value, c.fill = "R↓", f_renamed
                elif prev_present:
                    c.value, c.fill = "−", f_removed
                else:
                    c.value, c.fill = "", f_absent

    row_tot = len(dataset["all_fields"]) + 2
    ws.cell(row_tot, 1, "N° campi totali").font = Font(bold=True, size=9)
    for j, anno in enumerate(anni, 2):
        n = len(dataset["headers_per_anno"].get(anno, []))
        c = ws.cell(row_tot, j, n if n else "")
        c.font = Font(bold=True, size=9)
        c.alignment = Alignment(horizontal="center")

    row_leg = row_tot + 2
    ws.cell(row_leg, 1, "Legenda").font = Font(bold=True, size=9)
    for k, (label, fill) in enumerate([
        ("✓  Presente",                   f_present),
        ("+  Campo aggiunto",              f_new),
        ("−  Campo rimosso",               f_removed),
        ("R↑ / R↓  Potenziale rinomina",   f_renamed),
        ("   Assente / non disponibile",   f_absent),
    ], 1):
        c = ws.cell(row_leg + k, 1, label)
        c.fill = fill
        c.font = Font(size=9)

    ws.freeze_panes = "B2"


def write_variazioni_sheet(wb, all_datasets: list[dict], title: str = "Variazioni"):
    ws = wb.create_sheet(title=title)
    cols = ["Dataset", "Da anno", "A anno", "Tipo", "Campo", "→ Rinominato in",
            "N° campi da", "N° campi a", "Δ"]
    for j, h in enumerate(cols, 1):
        _hdr_cell(ws, 1, j, h)

    f_new     = PatternFill("solid", fgColor=C_NEW)
    f_removed = PatternFill("solid", fgColor=C_REMOVED)
    f_renamed = PatternFill("solid", fgColor=C_RENAMED)

    row = 2
    for ds in all_datasets:
        for v in ds["variazioni"]:
            renamed_new_set = set(v["rinominati"].values())
            renamed_old_set = set(v["rinominati"].keys())

            for campo in v["aggiunti"]:
                if campo in renamed_new_set:
                    continue
                ws.cell(row, 1, ds["nome"]); ws.cell(row, 2, v["da"]); ws.cell(row, 3, v["a"])
                c = ws.cell(row, 4, "AGGIUNTO"); c.fill = f_new; c.font = Font(bold=True, size=9)
                ws.cell(row, 5, campo); ws.cell(row, 7, v["n_da"])
                ws.cell(row, 8, v["n_a"]); ws.cell(row, 9, v["n_a"] - v["n_da"])
                row += 1

            for campo in v["rimossi"]:
                if campo in renamed_old_set:
                    continue
                ws.cell(row, 1, ds["nome"]); ws.cell(row, 2, v["da"]); ws.cell(row, 3, v["a"])
                c = ws.cell(row, 4, "RIMOSSO"); c.fill = f_removed; c.font = Font(bold=True, size=9)
                ws.cell(row, 5, campo); ws.cell(row, 7, v["n_da"])
                ws.cell(row, 8, v["n_a"]); ws.cell(row, 9, v["n_a"] - v["n_da"])
                row += 1

            for old, new in v["rinominati"].items():
                ws.cell(row, 1, ds["nome"]); ws.cell(row, 2, v["da"]); ws.cell(row, 3, v["a"])
                c = ws.cell(row, 4, "RINOMINATO?"); c.fill = f_renamed; c.font = Font(bold=True, size=9)
                ws.cell(row, 5, old); ws.cell(row, 6, new); ws.cell(row, 7, v["n_da"])
                ws.cell(row, 8, v["n_a"]); ws.cell(row, 9, v["n_a"] - v["n_da"])
                row += 1

    for col, w in zip("ABCDEFGHI", [28, 9, 9, 14, 40, 40, 11, 11, 6]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"


def write_confronto_sheet(wb, datasets: list[dict], anni: list[int], title: str):
    """Affianca tutte le categorie per anno (stessa struttura per ammessi e esclusi)."""
    ws = wb.create_sheet(title=title[:31])
    f_present = PatternFill("solid", fgColor=C_PRESENT)
    f_absent  = PatternFill("solid", fgColor=C_ABSENT)
    align_c   = Alignment(horizontal="center", vertical="center")

    row = 1
    for anno in sorted(anni, reverse=True):
        cat_con_dati = [ds for ds in datasets if ds["headers_per_anno"].get(anno)]
        if not cat_con_dati:
            continue

        c = ws.cell(row, 1, f"── Anno {anno} ──")
        c.font = Font(bold=True, size=11, color=C_HEADER)
        row += 1

        _hdr_cell(ws, row, 1, "Campo")
        ws.column_dimensions["A"].width = 42
        for j, ds in enumerate(cat_con_dati, 2):
            _hdr_cell(ws, row, j, ds["nome"])
            ws.column_dimensions[get_column_letter(j)].width = 18
        row += 1

        all_fields: list[str] = []
        seen: set[str] = set()
        for ds in cat_con_dati:
            for h in ds["headers_per_anno"].get(anno, []):
                if h not in seen:
                    all_fields.append(h)
                    seen.add(h)

        for field in all_fields:
            ws.cell(row, 1, field).font = Font(size=9)
            for j, ds in enumerate(cat_con_dati, 2):
                present = field in ds["headers_per_anno"].get(anno, [])
                c = ws.cell(row, j)
                c.value = "✓" if present else "—"
                c.fill = f_present if present else f_absent
                c.alignment = align_c
                c.font = Font(size=9)
            row += 1

        row += 2

    ws.freeze_panes = "B2"


# ---------------------------------------------------------------------------
# Raccolta dati
# ---------------------------------------------------------------------------

def raccogli_generale(dati_dir: Path, anni: list[int]) -> list[dict]:
    print("=== Elenco generale ===")
    headers: dict[int, list[str]] = {}
    for anno in anni:
        f = dati_dir / f"dati_{anno}.xlsx"
        if f.exists():
            h = get_headers(f)
            if h:
                headers[anno] = h
                print(f"  {anno}: {len(h)} campi")
            else:
                print(f"  {anno}: vuoto/illeggibile")
        else:
            print(f"  {anno}: mancante")

    if headers:
        return [analizza_dataset("Elenco_Generale", headers, anni)]
    return []


def raccogli_categorie(dati_dir: Path, anni: list[int], slugs: list[str]) -> dict[str, dict[str, dict]]:
    """
    Ritorna {slug: {"ammessi": dataset, "esclusi": dataset}}.
    """
    print("=== Categorie ===")
    risultati: dict[str, dict[str, dict]] = {}

    for slug in slugs:
        cat_dir = dati_dir / slug
        if not cat_dir.exists():
            print(f"  {slug}: cartella mancante")
            continue

        per_stato: dict[str, dict] = {}
        for stato, pattern in [("ammessi", "*_ammessi.xlsx"), ("esclusi", "*_esclusi.xlsx")]:
            headers: dict[int, list[str]] = {}
            for anno in anni:
                files = sorted(cat_dir.glob(f"{anno}_{pattern}"))
                if files:
                    h = get_headers(files[0])
                    if h:
                        headers[anno] = h

            if headers:
                nome = f"{slug}_{stato}"
                ds = analizza_dataset(nome, headers, anni)
                per_stato[stato] = ds
                anni_trovati = sorted(headers.keys())
                print(f"  {slug}/{stato}: {anni_trovati} — {[len(headers[a]) for a in anni_trovati]} campi")
            else:
                print(f"  {slug}/{stato}: nessun file trovato")

        if per_stato:
            risultati[slug] = per_stato

    return risultati

# ---------------------------------------------------------------------------
# Costruzione report
# ---------------------------------------------------------------------------

def build_report_generale(dati_dir: Path, anni: list[int], output: Path):
    datasets = raccogli_generale(dati_dir, anni)
    if not datasets:
        print("Nessun dato generale trovato.")
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for ds in datasets:
        write_matrix_sheet(wb, ds, anni)
        print(f"  Foglio '{ds['nome']}': {len(ds['all_fields'])} campi unici")

    write_variazioni_sheet(wb, datasets)
    print("  Foglio 'Variazioni'")

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    print(f"Report generale salvato: {output}")


def build_report_categorie(dati_dir: Path, anni: list[int], output: Path, slugs: list[str]):
    cat_data = raccogli_categorie(dati_dir, anni, slugs)
    if not cat_data:
        print("Nessun dato categorie trovato.")
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    all_ds: list[dict] = []
    ammessi_ds: list[dict] = []
    esclusi_ds: list[dict] = []

    for slug in slugs:
        if slug not in cat_data:
            continue
        for stato in ("ammessi", "esclusi"):
            if stato not in cat_data[slug]:
                continue
            ds = cat_data[slug][stato]
            write_matrix_sheet(wb, ds, anni)
            print(f"  Foglio '{ds['nome']}': {len(ds['all_fields'])} campi unici")
            all_ds.append(ds)
            if stato == "ammessi":
                ammessi_ds.append(ds)
            else:
                esclusi_ds.append(ds)

    write_variazioni_sheet(wb, all_ds, title="Variazioni")
    print("  Foglio 'Variazioni'")

    if len(ammessi_ds) > 1:
        write_confronto_sheet(wb, ammessi_ds, anni, "Confronto_Ammessi")
        print("  Foglio 'Confronto_Ammessi'")

    if len(esclusi_ds) > 1:
        write_confronto_sheet(wb, esclusi_ds, anni, "Confronto_Esclusi")
        print("  Foglio 'Confronto_Esclusi'")

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    print(f"Report categorie salvato: {output}")

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(
        description="Analisi evoluzione tracciati 5x1000",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Esempi:
  python analizza_tracciati.py                              # entrambi i report, ultimi 5 anni
  python analizza_tracciati.py --tipo generale
  python analizza_tracciati.py --tipo categorie
  python analizza_tracciati.py --tipo categorie --categoria ETS_ONLUS
  python analizza_tracciati.py --anni 3 --output-dir Dati/
        """,
    )
    ap.add_argument("--dati",       default="Dati",
                    help="Cartella Dati/ (default: Dati)")
    ap.add_argument("--anni",       type=int, default=5,
                    help="Ultimi N anni da analizzare (default: 5)")
    ap.add_argument("--tipo",       choices=["generale", "categorie", "entrambi"],
                    default="entrambi",
                    help="Quale report produrre (default: entrambi)")
    ap.add_argument("--categoria",  default="tutte",
                    help="Slug categoria specifica o 'tutte' (default: tutte). "
                         f"Valori: {', '.join(CATEGORIE_SLUG)}")
    ap.add_argument("--output-dir", default=None,
                    help="Cartella output (default: stessa di --dati)")
    ap.add_argument("--output",     default=None,
                    help="File output (solo per --tipo generale o categorie, "
                         "sovrascrive --output-dir)")
    args = ap.parse_args()

    dati_dir  = Path(args.dati)
    out_dir   = Path(args.output_dir) if args.output_dir else dati_dir

    if not dati_dir.exists():
        print(f"Errore: cartella non trovata: {dati_dir}")
        sys.exit(1)

    # Risolvi categoria
    if args.categoria.lower() == "tutte":
        slugs = CATEGORIE_SLUG
    else:
        # Cerca match esatto o prefix-insensitive
        match = next(
            (s for s in CATEGORIE_SLUG if s.lower() == args.categoria.lower()), None
        ) or next(
            (s for s in CATEGORIE_SLUG if args.categoria.lower() in s.lower()), None
        )
        if not match:
            print(f"Categoria non riconosciuta: {args.categoria}")
            print(f"Disponibili: {', '.join(CATEGORIE_SLUG)}")
            sys.exit(1)
        slugs = [match]

    # Anni disponibili
    anni_disponibili = sorted(
        int(f.stem.split("_")[1])
        for f in dati_dir.glob("dati_[0-9]*.xlsx")
        if len(f.stem.split("_")) >= 2 and f.stem.split("_")[1].isdigit()
    )
    if not anni_disponibili:
        print(f"Nessun file dati_XXXX.xlsx trovato in {dati_dir}/")
        sys.exit(1)

    anni = anni_disponibili[-args.anni:]
    print(f"Anni analizzati: {anni}")
    print()

    # Output paths
    if args.output and args.tipo != "entrambi":
        out_generale   = Path(args.output)
        out_categorie  = Path(args.output)
    else:
        suffix_cat = f"_{slugs[0]}" if len(slugs) == 1 else ""
        out_generale  = out_dir / "report_tracciati_generale.xlsx"
        out_categorie = out_dir / f"report_tracciati_categorie{suffix_cat}.xlsx"

    if args.tipo in ("generale", "entrambi"):
        print()
        build_report_generale(dati_dir, anni, out_generale)

    if args.tipo in ("categorie", "entrambi"):
        print()
        build_report_categorie(dati_dir, anni, out_categorie, slugs)


if __name__ == "__main__":
    main()
