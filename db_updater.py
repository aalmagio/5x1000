"""
db_updater.py
=============
Aggiorna il database MySQL del sito dopo ogni esecuzione della pipeline.

Configurazione tramite variabili d'ambiente (o file .env):
  SITE_DB_HOST      default: localhost
  SITE_DB_PORT      default: 3306
  SITE_DB_USER      obbligatorio
  SITE_DB_PASSWORD  obbligatorio
  SITE_DB_NAME      obbligatorio

Uso dalla pipeline:
    from db_updater import aggiorna_db_sito
    aggiorna_db_sito(
        anni_processati=[2023, 2024],
        steps_eseguiti=["download", "etl", "report"],
        csv_path=Path("Dati/enti_5x1000_norm.csv"),
        dati_dir=Path("Dati"),
        status="ok",
        t_inizio=time.time(),
    )
"""

from __future__ import annotations

import hashlib
import json
import logging
import os
import time
from datetime import datetime
from pathlib import Path

from db import get_connection, db_available

logger = logging.getLogger(__name__)


def _db_config(override: dict | None = None) -> dict:
    """Compatibilità retroattiva: restituisce params DB da env vars.
    Preferire get_connection() / db_available() da db.py per nuovo codice.
    """
    import os as _os
    cfg = {
        "host":     _os.getenv("SITE_DB_HOST", "localhost"),
        "port":     int(_os.getenv("SITE_DB_PORT", "3306")),
        "user":     _os.getenv("SITE_DB_USER", ""),
        "password": _os.getenv("SITE_DB_PASSWORD", ""),
        "database": _os.getenv("SITE_DB_NAME", ""),
        "charset":  "utf8mb4",
        "autocommit": True,
    }
    if override:
        cfg.update(override)
    return cfg


# ---------------------------------------------------------------------------
# Funzione principale
# ---------------------------------------------------------------------------

def registra_file_output(
    path: "Path | str",
    anno: "int | None" = None,
    tipo: str = "report",
    categoria: "str | None" = None,
    db_config: "dict | None" = None,
) -> bool:
    """
    Registra (o aggiorna) un singolo file nel catalogo ``dataset_files``.

    Utile per script stand-alone (es. ``report.py``) che generano file
    senza passare dall'intera pipeline.

    Parametri
    ---------
    path       : percorso al file generato
    anno       : anno di riferimento (None = file multi-anno)
    tipo       : 'report' | 'normalizzato' | 'completo' | 'categoria'
    categoria  : slug categoria (solo se tipo='categoria')
    db_config  : override configurazione DB (vedi _db_config)
    """
    try:
        import pymysql
    except ImportError:
        logger.warning("db_updater: pymysql non installato – registrazione file saltata")
        return False

    cfg = _db_config(db_config)
    if not cfg["user"] or not cfg["database"]:
        logger.warning("db_updater: SITE_DB_USER/SITE_DB_NAME non configurati – saltato")
        return False

    path = Path(path)
    if not path.exists():
        logger.warning(f"db_updater: file non trovato: {path}")
        return False

    ext = path.suffix.lstrip(".")
    if ext not in ("csv", "xlsx"):
        logger.warning(f"db_updater: formato non supportato ({ext}) per {path.name}")
        return False

    try:
        conn = pymysql.connect(**cfg)
        with conn:
            cur = conn.cursor()
            _upsert_file(cur, anno, tipo, categoria, ext, path)
        logger.info(f"db_updater: file registrato nel catalogo: {path.name}")
        return True
    except Exception as exc:
        logger.error(f"db_updater: errore registrazione file – {exc}", exc_info=True)
        return False


def aggiorna_db_sito(
    anni_processati: list[int],
    steps_eseguiti: list[str],
    csv_path: Path | str,
    dati_dir: Path | str,
    status: str = "ok",
    note: str = "",
    t_inizio: float | None = None,
    db_config: dict | None = None,
) -> bool:
    """
    Aggiorna il database del sito dopo un run della pipeline.

    Parametri
    ---------
    anni_processati : lista anni elaborati, es. [2023, 2024]
    steps_eseguiti  : lista step eseguiti, es. ["download","etl","report"]
    csv_path        : Path al CSV normalizzato principale
    dati_dir        : Path alla cartella Dati/
    status          : "ok" | "error" | "parziale"
    note            : messaggio libero (es. errore parziale)
    t_inizio        : timestamp float (time.time()) inizio pipeline
    db_config       : override configurazione DB

    Restituisce True se completato senza errori.
    """
    try:
        import pymysql
    except ImportError:
        logger.warning("db_updater: pymysql non installato – aggiornamento DB saltato")
        return False

    try:
        import pandas as pd
    except ImportError:
        logger.warning("db_updater: pandas non disponibile – aggiornamento DB saltato")
        return False

    cfg = _db_config(db_config)
    if not cfg["user"] or not cfg["database"]:
        logger.warning(
            "db_updater: SITE_DB_USER / SITE_DB_NAME non configurati – saltato"
        )
        return False

    csv_path = Path(csv_path) if csv_path else None
    dati_dir = Path(dati_dir) if dati_dir else None
    t0 = time.time()

    try:
        conn = pymysql.connect(**cfg)
        with conn:
            cur = conn.cursor()

            # 1. Conta righe nel CSV senza caricarlo tutto in memoria
            righe_totali = 0
            if csv_path and csv_path.exists():
                with open(csv_path, "r", encoding="utf-8", errors="replace") as fh:
                    righe_totali = sum(1 for _ in fh) - 1  # sottrai header

            # 2. Registra il run nella tabella pipeline_runs
            durata = int(t0 - t_inizio) if t_inizio else None
            cur.execute(
                """
                INSERT INTO pipeline_runs
                  (run_at, anni_processati, steps_eseguiti, righe_totali,
                   status, note, durata_secondi)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    datetime.now(),
                    json.dumps([int(a) for a in anni_processati]),
                    json.dumps(list(steps_eseguiti)),
                    righe_totali,
                    status,
                    note or None,
                    durata,
                ),
            )
            run_id = cur.lastrowid
            logger.info(f"db_updater: pipeline_run #{run_id} registrato ({status})")

            # 3. Aggiorna il catalogo dei file scaricabili
            if dati_dir and dati_dir.exists():
                _aggiorna_files(cur, dati_dir)
                logger.info("db_updater: dataset_files aggiornati")

            # 4. Aggiorna la tabella enti se il CSV esiste.
            # Viene eseguito anche quando ETL è saltato (smart mode) ma
            # la tabella enti è vuota o si vuole un reimport completo.
            if csv_path and csv_path.exists():
                # Controlla se la tabella è vuota: in quel caso importa sempre
                cur.execute("SELECT COUNT(*) FROM enti")
                row = cur.fetchone()
                count_enti = row[0] if row else 0
                force = (count_enti == 0)

                if "etl" in steps_eseguiti or force:
                    if force and "etl" not in steps_eseguiti:
                        logger.info("db_updater: tabella enti vuota – reimport completo dal CSV")
                    _aggiorna_enti(cur, csv_path, pd, anni_processati)
                    logger.info(
                        f"db_updater: tabella enti aggiornata per anni "
                        f"{anni_processati if anni_processati else 'tutti'}"
                    )
                    # 4b. Aggiorna tabella runts con denominazioni canoniche
                    try:
                        _aggiorna_runts(cur, csv_path, pd)
                        logger.info("db_updater: tabella runts aggiornata")
                    except Exception as exc_r:
                        logger.warning(f"db_updater: aggiornamento runts saltato – {exc_r}")

                    # 4c. Ricalcola ripartizioni aggregate (per categoria/regione)
                    try:
                        _aggiorna_ripartizioni(cur, anni_processati)
                        logger.info("db_updater: tabella ripartizioni aggiornata")
                    except Exception as exc_rip:
                        logger.warning(f"db_updater: aggiornamento ripartizioni saltato – {exc_rip}")

            # 4d. Popola categoria_ammissioni dai file per-categoria
            # (include importo_ripartizione_sotto_soglia e importo_totale_erogabile dal 2019)
            if dati_dir and dati_dir.exists():
                try:
                    n_cat = _aggiorna_categoria_ammissioni(cur, dati_dir)
                    logger.info(f"db_updater: categoria_ammissioni: {n_cat} righe aggiornate")
                except Exception as exc_cat:
                    logger.warning(f"db_updater: categoria_ammissioni saltato – {exc_cat}")

        logger.info(f"db_updater: completato in {time.time() - t0:.1f}s")
        return True

    except Exception as exc:
        logger.error(f"db_updater: errore durante l'aggiornamento – {exc}", exc_info=True)
        return False


def aggiorna_anno_ciclo(
    anno: int,
    csv_path: Path | str,
    dati_dir: Path | str,
    db_config: dict | None = None,
) -> bool:
    """
    Aggiorna il DB per un singolo anno in modalità ciclo (sito sempre attivo).

    Differenze rispetto a aggiorna_db_sito():
    - Usa UPSERT (INSERT … ON DUPLICATE KEY UPDATE) invece di DELETE+INSERT
      → nessuna finestra di "buco" nella disponibilità dei dati.
    - Aggiorna solo la tabella `enti` per l'anno indicato.
    - Aggiorna `categoria_ammissioni` solo per i file dell'anno indicato.
    - NON registra pipeline_run, NON aggiorna ripartizioni/runts
      (quelle vengono ricalcolate alla fine del ciclo completo).

    Ritorna True se completato senza errori.
    """
    try:
        import pymysql
        import pandas as pd
    except ImportError as e:
        logger.warning(f"db_updater: dipendenza mancante ({e}) – aggiorna_anno_ciclo saltato")
        return False

    cfg      = _db_config(db_config)
    if not cfg["user"] or not cfg["database"]:
        logger.warning("db_updater: SITE_DB_* non configurati – saltato")
        return False

    csv_path  = Path(csv_path) if csv_path else None
    dati_dir  = Path(dati_dir) if dati_dir else None
    t0        = time.time()

    try:
        conn = pymysql.connect(**cfg)
        with conn:
            cur = conn.cursor()

            # Enti: UPSERT per l'anno
            if csv_path and csv_path.exists():
                _aggiorna_enti(cur, csv_path, pd, [anno], upsert=True)
                logger.info(f"db_updater: enti anno {anno} aggiornati (upsert)")

            # categoria_ammissioni: solo file di questo anno
            if dati_dir and dati_dir.exists():
                totale_cat = 0
                for cat_dir in sorted(dati_dir.iterdir()):
                    if not cat_dir.is_dir():
                        continue
                    cat_slug = _CAT_SLUG.get(cat_dir.name.upper(), cat_dir.name.lower())
                    for stato in ("ammesso", "escluso"):
                        pattern = f"{anno}_*_{'ammessi' if stato == 'ammesso' else 'esclusi'}.xlsx"
                        for xlsx in sorted(cat_dir.glob(pattern)):
                            if anno >= 2019:
                                righe, *_ = _leggi_xlsx_ripartizione(xlsx)
                                if not righe:
                                    continue
                                batch = [
                                    (anno, cat_slug,
                                     r["cf"], r.get("denom"), stato,
                                     r["numero_scelte"], r["imp_esp"], r["imp_gen"], r["imp_ver"],
                                     r["imp_sottos"] or None, r["imp_erog"] or None)
                                    for r in righe
                                ]
                            else:
                                righe = _leggi_xlsx_ammissioni(xlsx)
                                if not righe:
                                    continue
                                batch = [
                                    (anno, cat_slug,
                                     r["cf"], r["denom"], stato,
                                     r["n_scelte"], r["importo_espresso"],
                                     r["importo_generico"], r["importo_totale"],
                                     None, None)
                                    for r in righe
                                ]
                            cur.executemany(_CAT_AMM_UPSERT, batch)
                            totale_cat += len(batch)

                if totale_cat:
                    logger.info(f"db_updater: categoria_ammissioni anno {anno}: {totale_cat} righe")

        logger.info(f"db_updater: aggiorna_anno_ciclo({anno}) completato in {time.time()-t0:.1f}s")
        return True

    except Exception as exc:
        logger.error(f"db_updater: aggiorna_anno_ciclo({anno}) errore – {exc}", exc_info=True)
        return False


def finalizza_ciclo(
    anni_processati: list[int],
    csv_path: Path | str,
    dati_dir: Path | str,
    t_inizio: float | None = None,
    db_config: dict | None = None,
) -> bool:
    """
    Operazioni finali da eseguire una sola volta al termine del ciclo completo:
    - Ricalcola ripartizioni aggregate
    - Aggiorna tabella runts
    - Registra pipeline_run con status='ok'

    Ritorna True se completato senza errori.
    """
    try:
        import pymysql
        import pandas as pd
    except ImportError as e:
        logger.warning(f"db_updater: dipendenza mancante ({e}) – finalizza_ciclo saltato")
        return False

    cfg = _db_config(db_config)
    if not cfg["user"] or not cfg["database"]:
        return False

    csv_path = Path(csv_path) if csv_path else None
    dati_dir = Path(dati_dir) if dati_dir else None
    t0       = time.time()

    try:
        conn = pymysql.connect(**cfg)
        with conn:
            cur = conn.cursor()

            # Runts
            if csv_path and csv_path.exists():
                try:
                    _aggiorna_runts(cur, csv_path, pd)
                    logger.info("db_updater: finalizza_ciclo – runts aggiornati")
                except Exception as e:
                    logger.warning(f"db_updater: finalizza_ciclo – runts saltato: {e}")

            # Ripartizioni
            try:
                _aggiorna_ripartizioni(cur, anni_processati)
                logger.info("db_updater: finalizza_ciclo – ripartizioni aggiornate")
            except Exception as e:
                logger.warning(f"db_updater: finalizza_ciclo – ripartizioni saltato: {e}")

            # File catalogo
            if dati_dir and dati_dir.exists():
                try:
                    _aggiorna_files(cur, dati_dir)
                except Exception as e:
                    logger.warning(f"db_updater: finalizza_ciclo – files saltato: {e}")

            # pipeline_run
            righe = 0
            if csv_path and csv_path.exists():
                with open(csv_path, "r", encoding="utf-8", errors="replace") as fh:
                    righe = sum(1 for _ in fh) - 1
            durata = int(time.time() - t_inizio) if t_inizio else None
            cur.execute(
                """INSERT INTO pipeline_runs
                   (run_at, anni_processati, steps_eseguiti, righe_totali, status, durata_secondi)
                   VALUES (%s, %s, %s, %s, 'ok', %s)""",
                (
                    datetime.now(),
                    json.dumps([int(a) for a in anni_processati]),
                    json.dumps(["ciclo"]),
                    righe,
                    durata,
                ),
            )

        logger.info(f"db_updater: finalizza_ciclo completato in {time.time()-t0:.1f}s")
        return True

    except Exception as exc:
        logger.error(f"db_updater: finalizza_ciclo errore – {exc}", exc_info=True)
        return False


# ---------------------------------------------------------------------------
# Aggiornamento catalogo file
# ---------------------------------------------------------------------------

# Mappa nomi directory categoria → slug per il DB
_CAT_SLUG = {
    "VOLONTARIATO": "volontariato",
    "ASD":          "asd",
    "ETS_ONLUS":    "ets_onlus",
    "RICERCA_SCI":  "ricerca_scientifica",
    "RICERCA_SAN":  "ricerca_sanitaria",
    "COMUNI":       "comuni",
    "BENI_CULT":    "beni_culturali",
    "AREE_PROT":    "aree_protette",
}


def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for block in iter(lambda: fh.read(65536), b""):
            h.update(block)
    return h.hexdigest()


def _upsert_file(cur, anno, tipo, categoria, formato, path: Path) -> None:
    if not path.exists():
        return
    size_mb = round(path.stat().st_size / 1_048_576, 2)
    sha = _sha256(path)
    cur.execute(
        """
        INSERT INTO dataset_files
          (anno, tipo, categoria, formato, percorso, nome_file,
           dimensione_mb, sha256, aggiornato_il)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW())
        ON DUPLICATE KEY UPDATE
          percorso       = VALUES(percorso),
          nome_file      = VALUES(nome_file),
          dimensione_mb  = VALUES(dimensione_mb),
          sha256         = VALUES(sha256),
          aggiornato_il  = NOW()
        """,
        (anno, tipo, categoria, formato, str(path), path.name, size_mb, sha),
    )


def _aggiorna_files(cur, dati_dir: Path) -> None:
    # Dataset completo (CSV)
    _upsert_file(cur, None, "completo", None, "csv", dati_dir / "enti_5x1000_norm.csv")
    _upsert_file(cur, None, "completo", None, "xlsx", dati_dir / "enti_5x1000_norm.xlsx")

    # File per anno (dati_YYYY.xlsx)
    for f in sorted(dati_dir.glob("dati_*.xlsx")):
        try:
            anno = int(f.stem.split("_")[1])
        except (IndexError, ValueError):
            continue
        _upsert_file(cur, anno, "normalizzato", None, "xlsx", f)

    # Report per anno (report_YYYY.xlsx)
    for f in sorted(dati_dir.glob("report_*.xlsx")):
        try:
            anno = int(f.stem.split("_")[1])
        except (IndexError, ValueError):
            continue
        _upsert_file(cur, anno, "report", None, "xlsx", f)

    # Pivot scelte per categoria (multi-anno)
    _upsert_file(cur, None, "scelte", None, "xlsx", dati_dir / "scelte_categorie.xlsx")

    # File per categoria + anno
    for cat_dir in sorted(dati_dir.iterdir()):
        if not cat_dir.is_dir():
            continue
        cat_slug = _CAT_SLUG.get(cat_dir.name.upper(), cat_dir.name.lower())
        for f in sorted(cat_dir.glob("*_ammessi.xlsx")):
            try:
                anno = int(f.stem.split("_")[0])
            except (IndexError, ValueError):
                continue
            _upsert_file(cur, anno, "categoria", cat_slug, "xlsx", f)


# ---------------------------------------------------------------------------
# Aggiornamento tabella categoria_ammissioni
# Legge i file *_ammessi.xlsx e *_esclusi.xlsx da ogni sottocartella di
# dati_dir e popola la tabella con (anno, categoria, cod_fiscale, stato).
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Lettura file categoria: CF, denominazione + metriche
# ---------------------------------------------------------------------------

# Alias colonna codice fiscale (case-insensitive, strip)
_CF_ALIASES    = {"codice fiscale", "cf", "codice_fiscale", "cod. fiscale", "c.f."}
# Alias colonna denominazione
_DENOM_ALIASES = {"denominazione", "beneficiario", "ente", "nome"}
# Alias metriche (cerchiamo prima match esatto, poi parziale)
_N_SCELTE_CAT_ALIASES = frozenset({
    "numero scelte", "n. scelte", "n scelte", "numero di scelte",
    "preferenze", "n preferenze", "numero preferenze",
})
_IMP_ESP_CAT_ALIASES = frozenset({
    "importo per scelte espresse", "importo scelte espresse",
    "importo (euro) per scelte espresse",
    "importo (euro) importo delle scelte espresse",
    "importo delle scelte espresse",
})
_IMP_GEN_CAT_ALIASES = frozenset({
    "importo proporzionale per le scelte generiche",
    "per scelte generiche", "importo generiche",
    "importo proporzionale scelte generiche",
})
_IMP_TOT_CAT_ALIASES = frozenset({
    "importo totale", "totale", "importo totale erogabile",
    "importo totale (euro)", "importo totale euro",
})


def _find_col_idx(header_norm: list[str], aliases: frozenset[str]) -> int | None:
    """Ritorna l'indice della colonna che matcha un alias (esatto poi parziale)."""
    for i, c in enumerate(header_norm):
        if c in aliases:
            return i
    for i, c in enumerate(header_norm):
        for a in aliases:
            if len(a) > 6 and (a in c or c in a):
                return i
    return None


def _parse_num(val) -> float | None:
    """Converte un valore grezzo di cella in float, None se non parsabile."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val) if val == val else None   # NaN check
    s = str(val).strip().replace(".", "").replace(",", ".").replace(" ", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _leggi_xlsx_ammissioni(path: Path) -> list[dict]:
    """
    Legge un file xlsx categoria e restituisce lista di dict con:
      cf, denom, n_scelte, importo_espresso, importo_generico, importo_totale

    Applica deduplicazione per CF: se lo stesso CF compare più volte,
    mantiene la prima denominazione non-null e somma le metriche numeriche.
    """
    try:
        import openpyxl
    except ImportError:
        logger.warning("db_updater: openpyxl non disponibile – categoria_ammissioni saltata")
        return []

    import warnings
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", "Workbook contains no default style", UserWarning)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []

    # Header normalizzato
    raw_header  = [str(c).strip() if c is not None else "" for c in rows[0]]
    header_norm = [h.lower() for h in raw_header]

    # Individua colonne
    cf_idx    = None
    denom_idx = None
    for i, c in enumerate(header_norm):
        if cf_idx    is None and c in _CF_ALIASES:
            cf_idx    = i
        if denom_idx is None and c in _DENOM_ALIASES:
            denom_idx = i
    # Fallback CF: "fiscal" o "codice"
    if cf_idx is None:
        for i, c in enumerate(header_norm):
            if "fiscal" in c or "codice" in c:
                cf_idx = i
                break
    if cf_idx is None:
        logger.warning(f"db_updater: colonna CF non trovata in {path.name} "
                       f"(header: {raw_header[:6]})")
        return []

    n_idx   = _find_col_idx(header_norm, _N_SCELTE_CAT_ALIASES)
    esp_idx = _find_col_idx(header_norm, _IMP_ESP_CAT_ALIASES)
    gen_idx = _find_col_idx(header_norm, _IMP_GEN_CAT_ALIASES)
    tot_idx = _find_col_idx(header_norm, _IMP_TOT_CAT_ALIASES)

    # Leggi righe dati
    seen: dict[str, dict] = {}
    for row in rows[1:]:
        if not row or cf_idx >= len(row):
            continue
        cf_raw = row[cf_idx]
        if cf_raw is None:
            continue
        cf = str(cf_raw).strip().upper()
        # Rimuovi artefatti ".0" da Excel (CF numerico letto come float)
        import re as _re
        cf = _re.sub(r"\.0+$", "", cf)
        if not cf:
            continue

        def _get(idx):
            return row[idx] if idx is not None and idx < len(row) else None

        denom = None
        if denom_idx is not None and denom_idx < len(row) and row[denom_idx] is not None:
            denom = str(row[denom_idx]).strip() or None

        n_scelte = _parse_num(_get(n_idx))
        imp_esp  = _parse_num(_get(esp_idx))
        imp_gen  = _parse_num(_get(gen_idx))
        imp_tot  = _parse_num(_get(tot_idx))
        # Intero per n_scelte
        n_scelte_int = int(n_scelte) if n_scelte is not None else None

        if cf not in seen:
            seen[cf] = {
                "cf":               cf,
                "denom":            denom,
                "n_scelte":         n_scelte_int,
                "importo_espresso": imp_esp,
                "importo_generico": imp_gen,
                "importo_totale":   imp_tot,
            }
        else:
            # CF duplicato: somma metriche, mantieni prima denom non-null
            e = seen[cf]
            if e["denom"] is None and denom is not None:
                e["denom"] = denom
            for k, v in (("n_scelte", n_scelte_int), ("importo_espresso", imp_esp),
                         ("importo_generico", imp_gen), ("importo_totale", imp_tot)):
                if v is not None:
                    e[k] = (e[k] or 0) + v

    return list(seen.values())


_CAT_AMM_UPSERT = """
    INSERT INTO categoria_ammissioni
      (anno, categoria, cod_fiscale, denominazione, stato,
       n_scelte, importo_espresso, importo_generico, importo_totale,
       importo_ripartizione_sotto_soglia, importo_totale_erogabile)
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    ON DUPLICATE KEY UPDATE
      denominazione                      = COALESCE(VALUES(denominazione), denominazione),
      stato                              = VALUES(stato),
      n_scelte                           = COALESCE(VALUES(n_scelte),         n_scelte),
      importo_espresso                   = COALESCE(VALUES(importo_espresso), importo_espresso),
      importo_generico                   = COALESCE(VALUES(importo_generico), importo_generico),
      importo_totale                     = COALESCE(VALUES(importo_totale),   importo_totale),
      importo_ripartizione_sotto_soglia  = COALESCE(VALUES(importo_ripartizione_sotto_soglia),
                                                    importo_ripartizione_sotto_soglia),
      importo_totale_erogabile           = COALESCE(VALUES(importo_totale_erogabile),
                                                    importo_totale_erogabile),
      aggiornato_il                      = NOW()
"""


def _aggiorna_categoria_ammissioni(cur, dati_dir: Path) -> int:
    """
    Popola `categoria_ammissioni` dai file *_ammessi.xlsx e *_esclusi.xlsx.

    - Anno < 2019: usa _leggi_xlsx_ammissioni (4 campi numerici,
      importo_ripartizione_sotto_soglia e importo_totale_erogabile = NULL).
    - Anno >= 2019: usa _leggi_xlsx_ripartizione (6 campi, include denom).
      importo_totale = importo_verifica_totale (espresso+generico).

    Aggiorna anche ripartizioni.n_scelte_generiche (totale nazionale per
    categoria/anno) estraendolo dal footer dei file *_ammessi.xlsx 2019+.

    Ritorna il numero totale di righe inserite/aggiornate.
    """
    totale = 0
    # (anno, cat_slug) → n_scelte_generiche estratto dal footer AdE
    gen_data: dict[tuple[int, str], int] = {}

    for cat_dir in sorted(dati_dir.iterdir()):
        if not cat_dir.is_dir():
            continue
        cat_slug = _CAT_SLUG.get(cat_dir.name.upper(), cat_dir.name.lower())

        for stato in ("ammesso", "escluso"):
            pattern = f"*_{'ammessi' if stato == 'ammesso' else 'esclusi'}.xlsx"
            for xlsx in sorted(cat_dir.glob(pattern)):
                try:
                    anno = int(xlsx.stem.split("_")[0])
                except (IndexError, ValueError):
                    logger.warning(f"db_updater: impossibile estrarre anno da {xlsx.name}")
                    continue

                if anno >= 2019:
                    righe, scartate, n_gen = _leggi_xlsx_ripartizione(xlsx)
                    if not righe:
                        continue
                    batch = [
                        (anno, cat_slug,
                         r["cf"], r.get("denom"), stato,
                         r["numero_scelte"], r["imp_esp"], r["imp_gen"], r["imp_ver"],
                         r["imp_sottos"] or None, r["imp_erog"] or None)
                        for r in righe
                    ]
                    if scartate:
                        logger.debug(f"db_updater: {xlsx.name}: {scartate} righe scartate")
                    # raccoglie n_scelte_generiche solo dal file ammessi (ha il footer)
                    if stato == "ammesso" and n_gen is not None:
                        gen_data[(anno, cat_slug)] = n_gen
                        logger.debug(
                            f"db_updater: {xlsx.name}: n_scelte_generiche={n_gen:,}"
                        )
                else:
                    righe = _leggi_xlsx_ammissioni(xlsx)
                    if not righe:
                        continue
                    batch = [
                        (anno, cat_slug,
                         r["cf"], r["denom"], stato,
                         r["n_scelte"], r["importo_espresso"],
                         r["importo_generico"], r["importo_totale"],
                         None, None)
                        for r in righe
                    ]

                cur.executemany(_CAT_AMM_UPSERT, batch)
                totale += len(batch)
                logger.info(
                    f"db_updater: categoria_ammissioni – {cat_slug}/{anno} "
                    f"{stato}: {len(batch)} righe"
                )

    # Aggiorna n_scelte_generiche nella riga nazionale di ripartizioni
    if gen_data:
        for (anno, cat), n_gen in gen_data.items():
            cur.execute(
                """UPDATE ripartizioni
                   SET n_scelte_generiche = %s
                   WHERE anno = %s AND categoria = %s AND regione IS NULL""",
                (n_gen, anno, cat),
            )
        logger.info(
            f"db_updater: n_scelte_generiche aggiornato per "
            f"{len(gen_data)} coppie (anno, categoria)"
        )

    return totale


# (categoria_ripartizioni rimossa: i campi sotto_soglia e totale_erogabile
#  sono ora colonne di categoria_ammissioni — vedi db_schema.sql)

# Alias per la quota redistribuita agli enti sotto soglia
_IMP_SOTTOS_ALIASES = frozenset({
    "importo ripartizione importi inferiori alla soglia",
    "importo ripartizione importi inferiori soglia",
    "importo ripartizione sotto soglia",
    "ripartizione importi inferiori alla soglia",
    "ripartizione sotto soglia",
    "redistribuzione",
    "ripartizione",
    "importo redistribuzione",
    "quota sotto soglia",
})

# Alias per importo totale erogabile (colonna finale, distinta da "importo totale")
_IMP_EROG_ALIASES = frozenset({
    "importo totale erogabile",
    "totale erogabile",
    "importo erogabile",
    "erogabile",
    "importo da erogare",
    "totale da erogare",
})

# Alias per la colonna composta espresso+generico (presente in alcuni file)
_IMP_VERIFICA_ALIASES = frozenset({
    "importo delle scelte espresse + importo proporzionale delle scelte generiche",
    "importo scelte espresse + importo proporzionale scelte generiche",
    "importo espresso + generico",
    "totale (espresso + generico)",
    "totale espresso generico",
    "importo totale (espresso + generico)",
})



def _leggi_xlsx_ripartizione(path: Path) -> tuple[list[dict], int, int | None]:
    """
    Legge un file xlsx categoria e restituisce (records, n_scartate, n_scelte_generiche).

    Ogni record è un dict con:
      cf, numero_scelte, imp_esp, imp_gen, imp_ver, imp_sottos, imp_erog

    n_scelte_generiche: contribuenti che hanno scelto la categoria senza indicare
    un ente specifico (riga "Voti generici" nel footer AdE). None se non trovata.

    Tutti i campi numerici mancanti vengono impostati a 0.
    importo_verifica_totale = colonna esplicita se presente, altrimenti
    somma di espresso+generico, altrimenti 0.

    Righe senza CF valido (len < 5) vengono scartate e contate.
    CF duplicati nello stesso file vengono deduplicati sommando i valori.
    """
    try:
        import openpyxl
    except ImportError:
        logger.warning("db_updater: openpyxl non disponibile")
        return [], 0

    import warnings
    import re as _re

    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", "Workbook contains no default style", UserWarning)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], 0, None

    raw_header  = [str(c).strip() if c is not None else "" for c in rows[0]]
    header_norm = [h.lower() for h in raw_header]

    # Colonna CF (stesso fallback già usato in _leggi_xlsx_ammissioni)
    cf_idx = None
    for i, c in enumerate(header_norm):
        if c in _CF_ALIASES:
            cf_idx = i
            break
    if cf_idx is None:
        for i, c in enumerate(header_norm):
            if "fiscal" in c or "codice" in c:
                cf_idx = i
                break
    if cf_idx is None:
        logger.warning(
            f"db_updater: CF non trovato in {path.name} (header: {raw_header[:6]})"
        )
        return [], 0, None

    # Denominazione (opzionale — usata quando il file viene letto per categoria_ammissioni)
    denom_idx = None
    for i, c in enumerate(header_norm):
        if c in _DENOM_ALIASES:
            denom_idx = i
            break

    # Individua indici colonne numeriche
    n_idx      = _find_col_idx(header_norm, _N_SCELTE_CAT_ALIASES)
    esp_idx    = _find_col_idx(header_norm, _IMP_ESP_CAT_ALIASES)
    gen_idx    = _find_col_idx(header_norm, _IMP_GEN_CAT_ALIASES)
    ver_idx    = _find_col_idx(header_norm, _IMP_VERIFICA_ALIASES)
    sottos_idx = _find_col_idx(header_norm, _IMP_SOTTOS_ALIASES)
    erog_idx   = _find_col_idx(header_norm, _IMP_EROG_ALIASES)
    # Fallback: se "totale erogabile" non trovato, usa alias "importo totale" generico
    if erog_idx is None:
        erog_idx = _find_col_idx(header_norm, _IMP_TOT_CAT_ALIASES)

    seen: dict[str, dict] = {}
    scartate = 0
    n_scelte_generiche: int | None = None

    def _check_footer_gen(row) -> None:
        nonlocal n_scelte_generiche
        label = str(row[0]).strip().lower() if row and row[0] is not None else ""
        if "voti generici" in label or "scelte generiche" in label:
            val = _parse_num(row[n_idx] if n_idx is not None and n_idx < len(row) else None)
            if val is not None:
                n_scelte_generiche = int(val)

    for row in rows[1:]:
        if not row or cf_idx >= len(row):
            continue
        cf_raw = row[cf_idx]
        if cf_raw is None:
            _check_footer_gen(row)
            scartate += 1
            continue

        cf = _re.sub(r"\.0+$", "", str(cf_raw).strip().upper()).strip("'\"")
        if not cf or len(cf) < 5:
            _check_footer_gen(row)
            scartate += 1
            continue

        def _get(idx):
            return row[idx] if idx is not None and idx < len(row) else None

        n_s   = _parse_num(_get(n_idx))
        i_esp = _parse_num(_get(esp_idx))
        i_gen = _parse_num(_get(gen_idx))
        i_ver = _parse_num(_get(ver_idx))
        i_sts = _parse_num(_get(sottos_idx))
        i_ero = _parse_num(_get(erog_idx))

        # importo_verifica_totale: colonna esplicita > somma derivata > 0
        if i_ver is None:
            if i_esp is not None and i_gen is not None:
                i_ver = i_esp + i_gen
            elif i_esp is not None:
                i_ver = i_esp
            else:
                i_ver = 0.0

        denom_raw = row[denom_idx] if denom_idx is not None and denom_idx < len(row) else None
        denom = str(denom_raw).strip() if denom_raw not in (None, "") else None

        rec = {
            "cf":            cf,
            "denom":         denom,
            "numero_scelte": int(n_s) if n_s is not None else 0,
            "imp_esp":       i_esp or 0.0,
            "imp_gen":       i_gen or 0.0,
            "imp_ver":       i_ver or 0.0,
            "imp_sottos":    i_sts or 0.0,
            "imp_erog":      i_ero or 0.0,
        }

        if cf not in seen:
            seen[cf] = rec
        else:
            e = seen[cf]
            e["numero_scelte"] += rec["numero_scelte"]
            for k in ("imp_esp", "imp_gen", "imp_ver", "imp_sottos", "imp_erog"):
                e[k] += rec[k]

    return list(seen.values()), scartate, n_scelte_generiche


# ---------------------------------------------------------------------------
# Aggiornamento tabella runts (denominazione canonica per CF)
# ---------------------------------------------------------------------------

def _aggiorna_runts(cur, csv_path: Path, pd) -> None:
    """
    Popola / aggiorna la tabella `runts` con una riga per codice fiscale.
    Strategia denominazione:
      - Se l'ente ha RUNTS_DENOMINAZIONE in almeno un anno → usa quella.
      - Altrimenti → usa la DENOMINAZIONE dell'anno più recente.
    Dati RUNTS (sezione, sede, data_iscrizione) presi dall'anno più recente
    con dati RUNTS disponibili.
    """
    cols_needed = {
        "ANNO", "COD_FISCALE", "DENOMINAZIONE",
        "RUNTS_DENOMINAZIONE", "RUNTS_SEZIONE",
        "RUNTS_SEDE_COMUNE", "RUNTS_SEDE_PROV",
        "RUNTS_5X1000", "RUNTS_DATA_ISCRIZIONE",
    }

    chunks = pd.read_csv(
        csv_path,
        dtype=str,
        chunksize=10_000,
        low_memory=False,
        usecols=lambda c: c in cols_needed,
    )
    all_rows = pd.concat(list(chunks), ignore_index=True)

    if "ANNO" in all_rows.columns:
        all_rows["ANNO"] = pd.to_numeric(all_rows["ANNO"], errors="coerce")

    # Riempi NaN con None-compatibili
    all_rows = all_rows.where(all_rows.notna(), other=None)

    result: dict[str, tuple] = {}

    for cf, group in all_rows.groupby("COD_FISCALE"):
        if not cf or str(cf) == "nan":
            continue
        group = group.sort_values("ANNO", ascending=False)

        # Righe con dati RUNTS
        runts_mask = (
            group["RUNTS_DENOMINAZIONE"].notna()
            & (group["RUNTS_DENOMINAZIONE"].astype(str).str.strip() != "")
        )
        runts_rows = group[runts_mask]

        if not runts_rows.empty:
            best = runts_rows.iloc[0]
            denom       = best.get("RUNTS_DENOMINAZIONE") or group.iloc[0].get("DENOMINAZIONE")
            sezione     = best.get("RUNTS_SEZIONE")        or None
            sede_comune = best.get("RUNTS_SEDE_COMUNE")   or None
            sede_prov   = best.get("RUNTS_SEDE_PROV")     or None
            raw_attivo  = str(best.get("RUNTS_5X1000", "0") or "0").strip()
            attivo      = 1 if raw_attivo in ("1", "True", "true") else 0
            data_isc    = _parse_date_ita(best.get("RUNTS_DATA_ISCRIZIONE"))
        else:
            best        = group.iloc[0]
            denom       = best.get("DENOMINAZIONE") or None
            sezione = sede_comune = sede_prov = data_isc = None
            attivo  = 0

        # Converti NaN / "nan" a None
        def _clean(v):
            if v is None:
                return None
            s = str(v).strip()
            return None if s in ("nan", "None", "") else s

        result[str(cf)] = (
            _clean(denom), _clean(sezione), _clean(sede_comune),
            _clean(sede_prov), attivo, _clean(data_isc),
        )

    upsert_sql = """
        INSERT INTO runts
          (cod_fiscale, denominazione, sezione, sede_comune, sede_prov,
           attivo_5x1000, data_iscrizione)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE
          denominazione   = VALUES(denominazione),
          sezione         = VALUES(sezione),
          sede_comune     = VALUES(sede_comune),
          sede_prov       = VALUES(sede_prov),
          attivo_5x1000   = VALUES(attivo_5x1000),
          data_iscrizione = VALUES(data_iscrizione),
          aggiornato_il   = NOW()
    """
    rows = [(cf,) + vals for cf, vals in result.items()]
    for i in range(0, len(rows), _CHUNK_SIZE):
        cur.executemany(upsert_sql, rows[i : i + _CHUNK_SIZE])

    logger.info(f"db_updater: {len(rows):,} righe upsert in tabella runts")


# ---------------------------------------------------------------------------
# Aggiornamento RUNTS da file Excel standalone
# ---------------------------------------------------------------------------

# Possibili nomi di colonna nel file RUNTS (uppercase, auto-detect)
_RUNTS_COL_ALIASES: dict[str, list[str]] = {
    "cod_fiscale":     ["CODICE FISCALE", "COD_FISCALE", "COD. FISCALE", "CF", "CODFISCALE"],
    "denominazione":   ["DENOMINAZIONE", "NOME", "RAGIONE SOCIALE"],
    "sezione":         ["SEZIONE", "SEZIONE RUNTS"],
    "sede_comune":     ["SEDE COMUNE", "SEDE_COMUNE", "COMUNE SEDE", "COMUNE LEGALE", "COMUNE"],
    "sede_prov":       ["SEDE PROV", "SEDE_PROV", "PROV SEDE", "PROVINCIA", "PROV", "SIGLA"],
    "attivo_5x1000":   ["5X1000", "ABILITATO 5X1000", "CINQUE PER MILLE", "RUNTS_5X1000"],
    "data_iscrizione": ["DATA ISCRIZIONE", "DATA_ISCRIZIONE", "RUNTS_DATA_ISCRIZIONE",
                        "DATA PRIMA ISCRIZIONE", "DATA"],
}


def aggiorna_runts_da_xlsx(xlsx_path: Path, conn) -> bool:
    """
    Aggiorna le tabelle `runts` e i campi RUNTS in `enti` da un file Excel RUNTS standalone.

    Accetta file con colonne in vari formati (auto-detect tramite _RUNTS_COL_ALIASES).
    Operazioni:
      1. UPSERT in `runts` (una riga per CF)
      2. UPDATE dei campi runts_* in `enti` per ogni CF trovato (tutti gli anni)

    Uso CLI:
      python db_updater.py --runts /path/to/elenco_runts.xlsx
    """
    try:
        import pandas as pd
    except ImportError:
        logger.error("db_updater: pandas non disponibile")
        return False

    logger.info(f"db_updater: lettura file RUNTS {xlsx_path}")
    try:
        df = pd.read_excel(xlsx_path, dtype=str)
    except Exception as e:
        logger.error(f"db_updater: impossibile leggere {xlsx_path}: {e}")
        return False

    # Normalizza intestazioni: strip + uppercase
    df.columns = [str(c).strip().upper() for c in df.columns]
    logger.info(f"db_updater: colonne rilevate: {list(df.columns)}")

    # Risolvi mapping colonne
    col = {}
    for field, aliases in _RUNTS_COL_ALIASES.items():
        for alias in aliases:
            if alias in df.columns:
                col[field] = alias
                break

    if "cod_fiscale" not in col:
        logger.error(
            f"db_updater: colonna codice fiscale non trovata. "
            f"Colonne disponibili: {list(df.columns)}"
        )
        return False

    logger.info(f"db_updater: mapping colonne RUNTS → {col}")

    # Costruisci lista righe normalizzate
    records: list[dict] = []
    for _, row in df.iterrows():
        cf = str(row.get(col["cod_fiscale"], "") or "").strip()
        if not cf or cf.lower() in ("nan", "none", ""):
            continue

        def _get(field: str) -> "str | None":
            if field not in col:
                return None
            v = str(row.get(col[field], "") or "").strip()
            return None if v.lower() in ("nan", "none", "") else v

        attivo_raw = _get("attivo_5x1000") or "0"
        attivo = 1 if attivo_raw.lower() in ("1", "true", "si", "sì", "yes", "x") else 0

        records.append({
            "cf":             cf,
            "denominazione":  _get("denominazione"),
            "sezione":        _get("sezione"),
            "sede_comune":    _get("sede_comune"),
            "sede_prov":      _get("sede_prov"),
            "attivo":         attivo,
            "data_isc":       _parse_date_ita(_get("data_iscrizione")),
        })

    if not records:
        logger.warning("db_updater: nessuna riga valida nel file RUNTS")
        return False

    logger.info(f"db_updater: {len(records):,} enti RUNTS da aggiornare")

    upsert_runts = """
        INSERT INTO runts
          (cod_fiscale, denominazione, sezione, sede_comune, sede_prov,
           attivo_5x1000, data_iscrizione)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE
          denominazione   = VALUES(denominazione),
          sezione         = VALUES(sezione),
          sede_comune     = VALUES(sede_comune),
          sede_prov       = VALUES(sede_prov),
          attivo_5x1000   = VALUES(attivo_5x1000),
          data_iscrizione = VALUES(data_iscrizione),
          aggiornato_il   = NOW()
    """

    update_enti = """
        UPDATE enti SET
          runts_denominazione   = %s,
          runts_sezione         = %s,
          runts_sede_comune     = %s,
          runts_sede_prov       = %s,
          runts_5x1000          = %s,
          runts_data_iscrizione = %s
        WHERE cod_fiscale = %s
    """

    try:
        with conn.cursor() as cur:
            # 1. UPSERT tabella runts
            batch_runts = [
                (r["cf"], r["denominazione"], r["sezione"], r["sede_comune"],
                 r["sede_prov"], r["attivo"], r["data_isc"])
                for r in records
            ]
            for i in range(0, len(batch_runts), _CHUNK_SIZE):
                cur.executemany(upsert_runts, batch_runts[i:i + _CHUNK_SIZE])
            logger.info(f"db_updater: runts aggiornata ({len(batch_runts):,} righe)")

            # 2. UPDATE campi RUNTS in tabella enti (tutti gli anni per quel CF)
            batch_enti = [
                (r["denominazione"], r["sezione"], r["sede_comune"],
                 r["sede_prov"], r["attivo"], r["data_isc"], r["cf"])
                for r in records
            ]
            updated = 0
            for i in range(0, len(batch_enti), _CHUNK_SIZE):
                cur.executemany(update_enti, batch_enti[i:i + _CHUNK_SIZE])
                updated += cur.rowcount
            logger.info(f"db_updater: enti aggiornata — {updated:,} righe modificate")

        conn.commit()
        return True
    except Exception as e:
        conn.rollback()
        logger.error(f"db_updater: errore aggiornamento RUNTS: {e}")
        return False


# ---------------------------------------------------------------------------
# Aggiornamento tabella enti
# ---------------------------------------------------------------------------

# Mapping esatto colonne CSV uppercase → colonne DB lowercase
_COL_MAP = {
    "ANNO":                  "anno",
    "COD_FISCALE":           "cod_fiscale",
    "DENOMINAZIONE":         "denominazione",
    "REGIONE":               "regione",
    "PROVINCIA":             "provincia",
    "COMUNE":                "comune",
    "CAT_VOLONTARIATO":      "cat_volontariato",
    "CAT_ASD":               "cat_asd",
    "CAT_ETS_ONLUS":         "cat_ets_onlus",
    "CAT_RICERCA_SCI":       "cat_ricerca_sci",
    "CAT_RICERCA_SAN":       "cat_ricerca_san",
    "CAT_COMUNI":            "cat_comuni",
    "CAT_BENI_CULT":         "cat_beni_cult",
    "CAT_AREE_PROT":         "cat_aree_prot",
    "CATEGORIA_PRINCIPALE":  "categoria_principale",
    "N_SCELTE":              "n_scelte",
    "IMPORTO_ESPRESSO":      "importo_espresso",
    "IMPORTO_GENERICO":      "importo_generico",
    "IMPORTO_RIPARTIZIONE":  "importo_ripartizione",
    "IMPORTO_TOTALE":        "importo_totale",
    "RUNTS_DENOMINAZIONE":   "runts_denominazione",
    "RUNTS_SEZIONE":         "runts_sezione",
    "RUNTS_SEDE_COMUNE":     "runts_sede_comune",
    "RUNTS_SEDE_PROV":       "runts_sede_prov",
    "RUNTS_5X1000":          "runts_5x1000",
    "RUNTS_DATA_ISCRIZIONE": "runts_data_iscrizione",
}

_CHUNK_SIZE = 5_000  # righe per batch INSERT

# Colonne booleane (TINYINT 0/1) che pandas scrive nel CSV come "True"/"False".
# Devono essere convertite in 1/0 prima dell'INSERT in MySQL, altrimenti
# la stringa "True" viene interpretata come 0 da MySQL → tutti i flag a 0.
_BOOL_COLS = frozenset({
    "cat_volontariato", "cat_asd", "cat_ets_onlus",
    "cat_ricerca_sci",  "cat_ricerca_san",
    "cat_comuni",       "cat_beni_cult", "cat_aree_prot",
    "runts_5x1000",
})


def _bool_to_int(v) -> int:
    """Converte 'True'/'False'/1/0/NaN → 1/0."""
    if v is None:
        return 0


_DATE_FORMATS = ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%Y/%m/%d")

def _parse_date_ita(v) -> "str | None":
    """Converte date in vari formati (dd/mm/yyyy, ecc.) → 'yyyy-mm-dd' per MySQL DATE.
    Restituisce None se il valore è vuoto o non parsabile."""
    if v is None:
        return None
    s = str(v).strip()
    if not s or s in ("nan", "None", "NaT", "0000-00-00"):
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    logger.debug(f"db_updater: data non parsabile ignorata: {s!r}")
    return None
    s = str(v).strip().lower()
    return 1 if s in ("true", "1", "yes", "sì", "si") else 0


def _aggiorna_enti(cur, csv_path: Path, pd, anni: list[int], upsert: bool = False) -> None:
    """
    Aggiorna la tabella `enti` per gli anni indicati.

    upsert=False (default, modalità reset):
      DELETE righe degli anni → INSERT. Garantisce coerenza ma crea una
      breve finestra in cui le righe dell'anno sono assenti.

    upsert=True (modalità ciclo, sito sempre attivo):
      INSERT ... ON DUPLICATE KEY UPDATE — nessuna cancellazione, ogni riga
      viene aggiornata sul posto. Richiede UNIQUE KEY (anno, cod_fiscale).

    Se `anni` è vuota reimporta TUTTI gli anni presenti nel CSV.
    """
    anni_int = [int(a) for a in anni]
    tutti    = len(anni_int) == 0
    cols_db  = list(_COL_MAP.values())
    ph_row   = "(" + ",".join(["%s"] * len(cols_db)) + ")"

    # Sempre upsert: gestisce duplicati nel CSV sia in modalità ciclo
    # (nessun DELETE) sia in reimport (dopo DELETE, l'ultima riga vince).
    update_cols = [c for c in cols_db if c not in ("anno", "cod_fiscale")]
    update_clause = ", ".join(f"{c} = VALUES({c})" for c in update_cols)
    insert_sql = (
        f"INSERT INTO enti ({','.join(cols_db)}) VALUES {ph_row} "
        f"ON DUPLICATE KEY UPDATE {update_clause}"
    )
    if upsert:
        logger.info(
            f"db_updater: UPSERT enti anni "
            f"{'tutti' if tutti else anni_int} (sito attivo, nessun DELETE)"
        )
    else:
        if tutti:
            cur.execute("DELETE FROM enti")
            logger.info("db_updater: cancellate TUTTE le righe da tabella enti (reimport completo)")
        else:
            placeholders = ",".join(["%s"] * len(anni_int))
            cur.execute(f"DELETE FROM enti WHERE anno IN ({placeholders})", anni_int)
            logger.info(f"db_updater: cancellate righe anni {anni_int} da tabella enti")

    anni_str  = {str(a) for a in anni_int}
    total_ins = 0

    for chunk in pd.read_csv(
        csv_path,
        dtype=str,
        chunksize=_CHUNK_SIZE,
        low_memory=False,
        on_bad_lines="warn",
    ):
        # Rinomina colonne CSV → DB
        chunk = chunk.rename(columns=_COL_MAP)

        # Filtra solo gli anni richiesti (CSV letto con dtype=str → confronto stringa)
        if not tutti and "anno" in chunk.columns:
            chunk = chunk[chunk["anno"].isin(anni_str)]
        if chunk.empty:
            continue

        # Deduplicazione difensiva: il CSV può avere duplicati (anno, CF)
        # se l'ETL è stato eseguito più volte in append o se un ente compare
        # in più categorie con righe separate. Tieni l'ultima occorrenza.
        if "anno" in chunk.columns and "cod_fiscale" in chunk.columns:
            before = len(chunk)
            chunk = chunk.drop_duplicates(subset=["anno", "cod_fiscale"], keep="last")
            dropped = before - len(chunk)
            if dropped:
                logger.warning(
                    f"db_updater: rimossi {dropped} duplicati (anno, cod_fiscale) nel chunk"
                )

        # Assicura che ci siano tutte le colonne (riempi mancanti con None)
        for col in cols_db:
            if col not in chunk.columns:
                chunk[col] = None

        # Converti colonne booleane da "True"/"False" stringa a 1/0 intero.
        # Pandas scrive i bool nel CSV come stringhe; MySQL TINYINT non
        # riconosce "True" come 1 e lo converte silenziosamente a 0.
        for col in _BOOL_COLS:
            if col in chunk.columns:
                chunk[col] = chunk[col].apply(_bool_to_int)

        # Converti date italiane dd/mm/yyyy → yyyy-mm-dd per MySQL DATE.
        if "runts_data_iscrizione" in chunk.columns:
            chunk["runts_data_iscrizione"] = chunk["runts_data_iscrizione"].apply(_parse_date_ita)

        # Converti NaN → None ed estrai righe come tuple senza iterrows().
        # to_numpy(dtype=object) è ~50x più veloce di iterrows() su 1M righe.

        arr = chunk[cols_db].to_numpy(dtype=object)
        arr[pd.isna(chunk[cols_db]).to_numpy()] = None
        rows = list(map(tuple, arr))
        cur.executemany(insert_sql, rows)
        total_ins += len(rows)

    logger.info(f"db_updater: inserite {total_ins:,} righe in tabella enti")


# ---------------------------------------------------------------------------
# Aggiornamento tabella ripartizioni
# ---------------------------------------------------------------------------

_DDL_RIPARTIZIONI = """
CREATE TABLE IF NOT EXISTS `ripartizioni` (
  `id`                   INT           NOT NULL AUTO_INCREMENT,
  `anno`                 SMALLINT      NOT NULL,
  `categoria`            VARCHAR(50)   NOT NULL,
  `regione`              VARCHAR(100)           DEFAULT NULL
                         COMMENT 'NULL = totale nazionale',
  `n_enti`               INT           NOT NULL DEFAULT 0
                         COMMENT 'nr. enti beneficiari',
  `n_contribuenti`       INT           NOT NULL DEFAULT 0
                         COMMENT 'firme (scelte espresse)',
  `n_scelte_generiche`   INT                    DEFAULT NULL
                         COMMENT 'contribuenti che hanno scelto la categoria senza indicare un ente (solo riga nazionale)',
  `importo_espresso`     DECIMAL(15,2)          DEFAULT NULL,
  `importo_generico`     DECIMAL(15,2)          DEFAULT NULL,
  `importo_totale`       DECIMAL(15,2)          DEFAULT NULL,
  `aggiornato_il`        DATETIME      NOT NULL DEFAULT CURRENT_TIMESTAMP
                         ON UPDATE CURRENT_TIMESTAMP,
  PRIMARY KEY (`id`),
  UNIQUE KEY `uq_rip` (`anno`, `categoria`, `regione`),
  KEY `idx_rip_anno` (`anno`),
  KEY `idx_rip_cat`  (`anno`, `categoria`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;
"""

# Migration idempotente per DB già esistenti
_ALTER_RIPARTIZIONI_N_GEN = """
ALTER TABLE `ripartizioni`
  ADD COLUMN `n_scelte_generiche` INT DEFAULT NULL
    COMMENT 'contribuenti che hanno scelto la categoria senza indicare un ente (solo riga nazionale)'
    AFTER `n_contribuenti`
"""

_UPSERT_RIP = """
INSERT INTO ripartizioni
  (anno, categoria, regione, n_enti, n_contribuenti,
   importo_espresso, importo_generico, importo_totale)
VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
ON DUPLICATE KEY UPDATE
  n_enti           = VALUES(n_enti),
  n_contribuenti   = VALUES(n_contribuenti),
  importo_espresso = VALUES(importo_espresso),
  importo_generico = VALUES(importo_generico),
  importo_totale   = VALUES(importo_totale),
  aggiornato_il    = NOW()
"""


def _aggiorna_ripartizioni(cur, anni: list[int]) -> None:
    """
    Ricalcola e aggiorna la tabella ripartizioni a partire dalla tabella enti.
    Produce due livelli di aggregazione per ogni (anno, categoria):
      - per regione  (granularità geografica)
      - nazionale    (regione = NULL)
    Va chiamata DOPO _aggiorna_enti().
    """
    cur.execute(_DDL_RIPARTIZIONI)
    try:
        cur.execute(_ALTER_RIPARTIZIONI_N_GEN)
    except Exception as e:
        # 1060 = Duplicate column name: colonna già presente → ignorabile
        if "1060" not in str(e) and "Duplicate column" not in str(e):
            raise

    anni_int = [int(a) for a in anni]
    tutti    = len(anni_int) == 0

    if tutti:
        cur.execute("DELETE FROM ripartizioni")
        anni_filter_sql = ""
        params_base: list = []
    else:
        ph = ",".join(["%s"] * len(anni_int))
        cur.execute(f"DELETE FROM ripartizioni WHERE anno IN ({ph})", anni_int)
        anni_filter_sql = f"WHERE anno IN ({ph}) AND categoria_principale IS NOT NULL"
        params_base = anni_int[:]

    # ── Per regione ──────────────────────────────────────────────────────────
    sql_reg = f"""
        SELECT anno, categoria_principale AS categoria, regione,
               COUNT(*)              AS n_enti,
               COALESCE(SUM(n_scelte), 0)          AS n_contribuenti,
               SUM(importo_espresso)               AS importo_espresso,
               SUM(importo_generico)               AS importo_generico,
               SUM(importo_totale)                 AS importo_totale
        FROM enti
        {anni_filter_sql if tutti else 'WHERE anno IN (' + ','.join(['%s']*len(anni_int)) + ') AND categoria_principale IS NOT NULL'}
        GROUP BY anno, categoria_principale, regione
    """
    cur.execute(sql_reg, params_base)
    rows_reg = cur.fetchall()
    if rows_reg:
        cur.executemany(_UPSERT_RIP, rows_reg)

    # ── Totali nazionali (regione = NULL) ─────────────────────────────────────
    sql_naz = f"""
        SELECT anno, categoria_principale AS categoria, NULL AS regione,
               COUNT(*)              AS n_enti,
               COALESCE(SUM(n_scelte), 0)          AS n_contribuenti,
               SUM(importo_espresso)               AS importo_espresso,
               SUM(importo_generico)               AS importo_generico,
               SUM(importo_totale)                 AS importo_totale
        FROM enti
        {anni_filter_sql if tutti else 'WHERE anno IN (' + ','.join(['%s']*len(anni_int)) + ') AND categoria_principale IS NOT NULL'}
        GROUP BY anno, categoria_principale
    """
    cur.execute(sql_naz, params_base)
    rows_naz = cur.fetchall()
    if rows_naz:
        cur.executemany(_UPSERT_RIP, rows_naz)

    logger.info(
        f"db_updater: ripartizioni aggiornate — "
        f"{len(rows_reg):,} righe per regione, {len(rows_naz):,} nazionali"
    )


# ---------------------------------------------------------------------------
# Esecuzione diretta: python db_updater.py [--anni 2023,2024]
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import argparse, pathlib, sys

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[logging.StreamHandler(sys.stdout)],
    )

    # Carica .env dalla root del progetto (stesso comportamento della pipeline)
    _env_file = pathlib.Path(__file__).parent / ".env"
    if _env_file.is_file():
        with open(_env_file, encoding="utf-8") as _fh:
            for _line in _fh:
                _line = _line.strip()
                if not _line or _line.startswith("#") or "=" not in _line:
                    continue
                _k, _v = _line.split("=", 1)
                _k = _k.strip()
                _v = _v.strip().strip("\"'")
                if _k and not os.environ.get(_k):
                    os.environ[_k] = _v

    ap = argparse.ArgumentParser(description="Reimporta il CSV nel DB del sito")
    ap.add_argument("--anni", default=None,
                    help="Anni da reimportare, es. 2023,2024 (default: tutti)")
    ap.add_argument("--csv", default=None,
                    help="Percorso al CSV normalizzato (default: Dati/enti_5x1000_norm.csv)")
    ap.add_argument("--dati-dir", default=None,
                    help="Cartella Dati/ (default: ./Dati)")
    ap.add_argument("--runts", default=None, metavar="FILE.xlsx",
                    help="Aggiorna solo i dati RUNTS da un file Excel standalone "
                         "(aggiorna tabelle runts ed enti, salta tutto il resto)")
    ap.add_argument("--solo-categorie", action="store_true",
                    help="Aggiorna solo categoria_ammissioni e ripartizioni "
                         "(salta il reimport del CSV, veloce)")
    args = ap.parse_args()

    root = pathlib.Path(__file__).parent
    dati_dir = pathlib.Path(args.dati_dir) if args.dati_dir else root / "Dati"
    csv_path = pathlib.Path(args.csv) if args.csv else dati_dir / "enti_5x1000_norm.csv"

    # ── Modalità --solo-categorie: aggiorna categoria_ammissioni e ripartizioni ──
    if args.solo_categorie:
        import pymysql
        cfg = _db_config()
        if not cfg["user"] or not cfg["database"]:
            logger.error("SITE_DB_* non configurati — imposta le variabili d'ambiente")
            sys.exit(1)
        try:
            conn = pymysql.connect(**{**cfg, "autocommit": False})
        except Exception as e:
            logger.error(f"Impossibile connettersi al DB: {e}")
            sys.exit(1)
        try:
            with conn:
                cur = conn.cursor()
                n_cat = _aggiorna_categoria_ammissioni(cur, dati_dir)
                logger.info(f"db_updater: categoria_ammissioni: {n_cat} righe aggiornate")
                anni_list = []
                if args.anni:
                    try:
                        anni_list = [int(a.strip()) for a in args.anni.split(",") if a.strip()]
                    except ValueError:
                        logger.error("--anni deve contenere numeri, es. 2023,2024")
                        sys.exit(1)
                _aggiorna_ripartizioni(cur, anni_list)
                logger.info("db_updater: ripartizioni aggiornate")
        except Exception as e:
            logger.error(f"db_updater: errore – {e}", exc_info=True)
            sys.exit(1)
        sys.exit(0)

    # ── Modalità --runts: aggiorna solo dati RUNTS ───────────────────────────
    if args.runts:
        xlsx_path = pathlib.Path(args.runts)
        if not xlsx_path.is_file():
            logger.error(f"File non trovato: {xlsx_path}")
            sys.exit(1)
        cfg = _db_config()
        if not cfg["user"] or not cfg["database"]:
            logger.error("SITE_DB_* non configurati — imposta le variabili d'ambiente")
            sys.exit(1)
        try:
            conn = pymysql.connect(**{**cfg, "autocommit": False})
        except Exception as e:
            logger.error(f"Impossibile connettersi al DB: {e}")
            sys.exit(1)
        ok = aggiorna_runts_da_xlsx(xlsx_path, conn)
        conn.close()
        sys.exit(0 if ok else 1)

    # ── Modalità normale: reimport CSV completo ──────────────────────────────
    anni_list = []
    if args.anni:
        try:
            anni_list = [int(a.strip()) for a in args.anni.split(",") if a.strip()]
        except ValueError:
            logger.error("--anni deve contenere numeri, es. 2023,2024")
            sys.exit(1)

    ok = aggiorna_db_sito(
        anni_processati=anni_list,
        steps_eseguiti=["etl"],   # forza aggiornamento enti
        csv_path=csv_path,
        dati_dir=dati_dir,
        status="ok",
        note="Reimport manuale",
    )
    sys.exit(0 if ok else 1)
