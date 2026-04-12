#!/usr/bin/env python3
"""
5 per Mille - Download e conversione dati dall'Agenzia delle Entrate.

Scarica i file PDF/CSV dalle pagine ufficiali dell'Agenzia delle Entrate,
li organizza in cartelle per anno e li converte in file Excel.

Uso:
    python cinque_per_mille.py                                    # interattivo
    python cinque_per_mille.py --no-download --source csv         # batch: solo conversione CSV
    python cinque_per_mille.py --source csv --anni 2023,2024      # batch: anni specifici
    python cinque_per_mille.py --help                             # mostra tutti i flag

Requisiti:
    pip install -r requirements.txt
"""

import sys
import os
import re
import glob
import csv
import shutil
import logging
import datetime
import argparse
import pdfplumber
from io import StringIO
from urllib.parse import urljoin, urlparse, unquote
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from common import load_config, get_logger, ask_yes_no
from openpyxl.utils import get_column_letter

try:
    import requests
    from bs4 import BeautifulSoup
    HAS_WEB = True
except ImportError:
    HAS_WEB = False

# ============================================================================
# CONFIGURAZIONE
# ============================================================================

def _apply_config(cfg):
    """
    Applica i valori di config.yaml alle variabili globali di configurazione.
    I valori presenti nel file sovrascrivono i default; i valori assenti
    lasciano i default invariati.
    """
    global YEAR_URLS, HEADERS, TIMEOUT_PAGE, TIMEOUT_FILE

    # URL per anno
    url_anni = cfg.get("url_anni")
    if url_anni and isinstance(url_anni, dict):
        for anno, url in url_anni.items():
            YEAR_URLS[int(anno)] = str(url)

    # Download: headers e timeout
    dl_cfg = cfg.get("download", {})
    if dl_cfg.get("user_agent"):
        HEADERS["User-Agent"] = dl_cfg["user_agent"]
    if dl_cfg.get("timeout_pagina"):
        TIMEOUT_PAGE = int(dl_cfg["timeout_pagina"])
    if dl_cfg.get("timeout_file"):
        TIMEOUT_FILE = int(dl_cfg["timeout_file"])


YEAR_URLS = {
    2024: "https://www.agenziaentrate.gov.it/portale/elenco-complessivo-dei-beneficiari",
    2023: "https://www.agenziaentrate.gov.it/portale/elenco-complessivo-dei-beneficiari-2023",
    2022: "https://www.agenziaentrate.gov.it/portale/elenco-complessivo-degli-enti-ammessi-in-una-o-piu-categorie-di-beneficiari",
    2021: "https://www.agenziaentrate.gov.it/portale/elenco-complessivo-degli-enti-ammessi-in-una-o-pi%C3%B9-categorie-di-beneficiari",
    2020: "https://www.agenziaentrate.gov.it/portale/web/guest/elenco-5x1000-2020-enti-ammessi-categorie-di-beneficiari",
    2019: "https://www.agenziaentrate.gov.it/portale/elenco-complessivo-dei-beneficiari-2019",
    2018: "https://www.agenziaentrate.gov.it/portale/elenco-complessivo-beneficiari-2018",
    2017: "https://www.agenziaentrate.gov.it/portale/archivio/archivioschedeadempimento/schede-adempimento-2017/agevolazioni-2017/iscrizione-elenchi-5-per-mille-2017/elenchi-5xmille2017/elenco-completo-beneficiari-5xmille2017",
    2016: "https://www.agenziaentrate.gov.it/portale/archivio/archivioschedeadempimento/schede-adempimento-2016/richiedere-2016/iscrizione-elenchi-5-per-mille-2016/elenchi-5xmille2016/elenco-enti-ammessi-categorie-beneficiari",
    2015: "https://www.agenziaentrate.gov.it/portale/web/guest/archivio/archivioschedeadempimento/schede-adempimento-2015/richiedere-2015/iscrizione-elenchi-5-per-mille-2015/elenchi-5xmille2015/elenco-complessivo-beneficiari-5xmille2015",
    2014: "https://www.agenziaentrate.gov.it/portale/web/guest/archivio/archivioschedeadempimento/schede-adempimento-2014/richiedere-2014/iscrizione-elenchi-5-per-mille-2014/elenchi-5xmille2014/elenco-complessivo-beneficiari-5xmille2014",
    2013: "https://www.agenziaentrate.gov.it/portale/web/guest/archivio/archivioschedeadempimento/schede-adempimento-2013/richiedere-2013/contributo-5xmille2013/elenchi-5xmille2013/elenco-complessivo-beneficiari-5xmille2013",
    2012: "https://www.agenziaentrate.gov.it/portale/web/guest/archivio/archivioschedeadempimento/schede-adempimento-2012/richiedere-2012/contributo-5-per-mille-2012/elenchi-2012/elenco-complessivo-beneficiari",
    2011: "https://www.agenziaentrate.gov.it/portale/web/guest/archivio/archivioschedeadempimento/schede-adempimento-2011/richiedere-2011/contributo-del-5-per-mille-2011/elenchi-2011/elenco-complessivo-beneficiari-2011",
    2010: "https://www.agenziaentrate.gov.it/portale/web/guest/archivio/archivioschedeadempimento/schede-adempimento-2010/richiedere-2010/iscrizione-elenchi-5-per-mille/elenchi/elenco-complessivo-beneficiari-5permille2010",
    2009: "https://www.agenziaentrate.gov.it/portale/web/guest/archivio/archivio-5permille/5-per-mille-anni-precedenti/contributo-5-per-mille-2009/elenchi-5xmille2009/elenco-complessivo-beneficiari-5xmille2009",
    2008: "https://www.agenziaentrate.gov.it/portale/web/guest/archivio/archivio-5permille/5-per-mille-anni-precedenti/contributo-5-per-mille-2008/elenchi-5xmille2008/elenco-complessivo-beneficiari-5xmille2008",
    2007: "https://www.agenziaentrate.gov.it/portale/web/guest/archivio/archivio-5permille/5-per-mille-anni-precedenti/contributo-5-per-mille-2007/elenchi-5xmille2007/elenchi-complessivi-5xmille2007",
    2006: "https://www.agenziaentrate.gov.it/portale/web/guest/archivio/archivio-5permille/5-per-mille-anni-precedenti/contributo-5-per-mille-2006/elenchi-5xmille2006/elenco-complessivo-beneficiari-5xmille2006",
}

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "it-IT,it;q=0.9,en;q=0.8",
}

TIMEOUT_PAGE = 60       # Secondi di attesa per le pagine web
TIMEOUT_FILE = 120      # Secondi di attesa per il download file

# ============================================================================
# DOWNLOAD
# ============================================================================

def find_download_links(page_url, session):
    """
    Scarica la pagina e cerca tutti i link a file PDF e CSV.
    Restituisce dict {"pdf": [urls], "csv": [urls]}.
    """
    logging.info(f"  Scarico pagina: {page_url}")
    try:
        resp = session.get(page_url, headers=HEADERS, timeout=TIMEOUT_PAGE)
        resp.raise_for_status()
    except requests.RequestException as e:
        logging.error(f"  Errore nel scaricare la pagina: {e}")
        return {"pdf": [], "csv": []}

    soup = BeautifulSoup(resp.text, "html.parser")
    links = {"pdf": [], "csv": []}

    for a_tag in soup.find_all("a", href=True):
        href = a_tag["href"].strip()
        href_lower = href.lower()

        # Filtra solo link a file PDF o CSV
        is_pdf = href_lower.endswith(".pdf") or ".pdf" in href_lower
        is_csv = href_lower.endswith(".csv") or ".csv" in href_lower

        if not is_pdf and not is_csv:
            # Controlla anche il testo del link
            link_text = a_tag.get_text(strip=True).lower()
            if "pdf" in link_text and ("elenco" in link_text or "beneficiari" in link_text or "parte" in link_text):
                is_pdf = True
            elif "csv" in link_text and ("elenco" in link_text or "beneficiari" in link_text or "parte" in link_text):
                is_csv = True

        if is_pdf or is_csv:
            full_url = urljoin(page_url, href)
            file_type = "pdf" if is_pdf else "csv"
            if full_url not in links[file_type]:
                links[file_type].append(full_url)
                link_text = a_tag.get_text(strip=True)[:80]
                logging.info(f"    Trovato {file_type.upper()}: {link_text}")

    logging.info(f"  Totale link trovati: {len(links['pdf'])} PDF, {len(links['csv'])} CSV")
    return links


def download_file(url, dest_path, session):
    """Scarica un file da URL e lo salva in dest_path. Restituisce True se riuscito."""
    filename = os.path.basename(dest_path)
    if os.path.exists(dest_path):
        size = os.path.getsize(dest_path)
        logging.info(f"    File gia' presente ({size:,} bytes), salto: {filename}")
        return True

    logging.info(f"    Scarico: {filename}...")
    try:
        resp = session.get(url, headers=HEADERS, timeout=TIMEOUT_FILE, stream=True)
        resp.raise_for_status()

        with open(dest_path, "wb") as f:
            for chunk in resp.iter_content(chunk_size=8192):
                f.write(chunk)

        size = os.path.getsize(dest_path)
        logging.info(f"    Salvato: {filename} ({size:,} bytes)")
        return True
    except requests.RequestException as e:
        logging.error(f"    Errore nel download di {filename}: {e}")
        if os.path.exists(dest_path):
            os.remove(dest_path)
        return False


def sanitize_filename(url, index, ext):
    """Genera un nome file pulito dall'URL, con indice progressivo."""
    parsed = urlparse(url)
    basename = os.path.basename(unquote(parsed.path))
    if basename and len(basename) < 200:
        # Pulisci il nome
        basename = re.sub(r'[^\w\-\.\(\) ]', '_', basename)
        if not basename.lower().endswith(f".{ext}"):
            basename += f".{ext}"
        return basename
    return f"file_{index:02d}.{ext}"


def download_year(year, root_dir, session):
    """Scarica tutti i file per un anno. Restituisce il numero di file scaricati."""
    url = YEAR_URLS.get(year)
    if not url:
        logging.warning(f"[{year}] Nessun URL configurato")
        return 0

    logging.info(f"[{year}] Inizio download")
    folder = os.path.join(root_dir, str(year))
    os.makedirs(folder, exist_ok=True)

    links = find_download_links(url, session)
    total_pdf = len(links["pdf"])
    total_csv = len(links["csv"])

    if total_pdf == 0 and total_csv == 0:
        logging.warning(f"[{year}] Nessun file trovato nella pagina. Potrebbe essere necessario verificare manualmente.")
        return 0

    downloaded = 0
    for ext in ["pdf", "csv"]:
        for idx, file_url in enumerate(links[ext], 1):
            filename = sanitize_filename(file_url, idx, ext)
            dest = os.path.join(folder, filename)
            if download_file(file_url, dest, session):
                downloaded += 1

    logging.info(f"[{year}] Download completato: {downloaded} file scaricati in {folder}")
    return downloaded

# ============================================================================
# PDF EXTRACTION (dal vecchio script)
# ============================================================================

def clean_cell(value):
    if value is None:
        return ""
    value = str(value).strip()
    # Rimuovi apici che circondano il valore (artefatto CSV per forzare testo)
    if len(value) >= 2 and value[0] in ("'", '"') and value[-1] == value[0]:
        value = value[1:-1].strip()
    value = re.sub(r'\n+', ' ', value)
    value = re.sub(r'\s{2,}', ' ', value)
    value = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', value)
    # Compatta testo spaziato tipo "L O M B A R D I A" -> "LOMBARDIA"
    # Cerca sequenze di 3+ caratteri singoli separati da spazio
    value = re.sub(
        r'(?<![A-Za-z0-9])([A-Za-z0-9])((?:\s[A-Za-z0-9]){2,})(?![A-Za-z0-9])',
        lambda m: m.group(0).replace(' ', ''),
        value,
    )
    return value


def extract_header_from_pdf(pdf_path):
    """Estrae l'intestazione dal primo PDF."""
    try:
        pdf = pdfplumber.open(pdf_path)
        tables = pdf.pages[0].extract_tables()
        pdf.close()
    except Exception as e:
        logging.error(f"  Errore nel leggere intestazione da {pdf_path}: {e}")
        return None, None, 0

    if not tables or not tables[0]:
        return None, None, 0

    table = tables[0]
    row1 = table[0]
    num_cols = len(row1)

    # Verifica se c'e' una seconda riga di sotto-intestazione
    if len(table) >= 2 and table[1][0] is None:
        row2 = table[1]
        header = []
        for i in range(num_cols):
            top = clean_cell(row1[i]) if i < len(row1) else ""
            bottom = clean_cell(row2[i]) if i < len(row2) else ""
            if top and bottom:
                header.append(f"{top} - {bottom}")
            elif top:
                header.append(top)
            elif bottom:
                header.append(bottom)
            else:
                header.append(f"Colonna_{i+1}")
    else:
        header = [clean_cell(c) for c in row1]

    # Firma per riconoscere intestazioni ripetute
    signature = [c for c in [clean_cell(c) for c in row1] if c][:3]
    return header, signature, num_cols


def is_header_row(row, header_signature):
    cleaned = [clean_cell(c) for c in row]
    row_sig = [c for c in cleaned if c][:3]
    return row_sig == header_signature


def extract_data_from_pdf(pdf_path, header_signature, num_cols):
    """
    Generatore: yielda le righe del PDF una pagina alla volta.
    Chiude ogni pagina dopo l'elaborazione e richiama gc.collect() ogni
    20 pagine per contenere il consumo di RAM con PDF di grandi dimensioni.
    """
    import gc
    try:
        pdf = pdfplumber.open(pdf_path)
    except Exception as e:
        logging.error(f"  Errore nell'aprire {pdf_path}: {e}")
        return

    try:
        for page_idx, page in enumerate(pdf.pages):
            try:
                tables = page.extract_tables()
            except Exception as e:
                logging.warning(f"  Errore pagina {page_idx + 1} di {os.path.basename(pdf_path)}: {e}")
                continue

            for table in tables:
                for row in table:
                    cleaned = [clean_cell(c) for c in row]

                    if is_header_row(row, header_signature):
                        continue

                    if all(c == "" for c in cleaned):
                        continue

                    # Sotto-intestazione
                    if cleaned[0] == "" and any(c != "" for c in cleaned[4:]):
                        non_empty = [c for c in cleaned if c]
                        if all(len(c) < 40 for c in non_empty) and len(non_empty) <= num_cols // 2:
                            continue

                    while len(cleaned) < num_cols:
                        cleaned.append("")
                    cleaned = cleaned[:num_cols]
                    yield cleaned

            # Libera la pagina dalla cache di pdfplumber e forza il GC
            page.flush_cache()
            if page_idx % 20 == 0:
                gc.collect()
    finally:
        pdf.close()
        gc.collect()

# ============================================================================
# CSV EXTRACTION
# ============================================================================

def detect_csv_params(file_path):
    """Rileva delimitatore e encoding di un file CSV."""
    # Prova encoding
    for enc in ["utf-8", "latin-1", "cp1252", "iso-8859-15"]:
        try:
            with open(file_path, "r", encoding=enc) as f:
                sample = f.read(4096)
            break
        except (UnicodeDecodeError, UnicodeError):
            continue
    else:
        enc = "latin-1"
        with open(file_path, "r", encoding=enc) as f:
            sample = f.read(4096)

    # Rileva delimitatore
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ";"  # Default per dati italiani

    return enc, delimiter


def extract_data_from_csv(csv_path):
    """
    Legge un file CSV e restituisce (header, rows).
    Gestisce i vari formati dell'Agenzia delle Entrate.
    """
    enc, delimiter = detect_csv_params(csv_path)
    logging.info(f"    CSV: encoding={enc}, delimitatore='{delimiter}'")

    rows = []
    header = None
    title_checked = False

    with open(csv_path, "r", encoding=enc, errors="replace") as f:
        reader = csv.reader(f, delimiter=delimiter)
        for row_idx, row in enumerate(reader):
            cleaned = [clean_cell(c) for c in row]

            # Salta righe vuote
            if all(c == "" for c in cleaned):
                continue

            # Prima riga non vuota = intestazione (candidata)
            if header is None:
                header = cleaned
                continue

            # Controlla se l'header era in realta' una riga di titolo:
            # se ha meno colonne non-vuote dei dati, la vera intestazione
            # e' questa riga corrente.
            if not title_checked:
                title_checked = True
                hdr_filled = sum(1 for c in header if c)
                row_filled = sum(1 for c in cleaned if c)
                if hdr_filled < row_filled:
                    header = cleaned
                    continue

            # Salta intestazioni ripetute (stessa prima cella dell'header)
            if cleaned[:3] == header[:3]:
                continue

            rows.append(cleaned)

    return header, rows

# ============================================================================
# EXCEL CREATION
# ============================================================================

# Configurazione Excel di default (sovrascritta da config.yaml)
EXCEL_CONFIG = {
    "header_font": "Arial",
    "header_size": 10,
    "header_bg": "2F5496",
    "header_fg": "FFFFFF",
    "data_font": "Arial",
    "data_size": 10,
    "alt_row_bg": "EEF2FF",
    "border_color": "D9D9D9",
    "max_col_width": 45,
}


def _apply_excel_config(cfg):
    """Aggiorna EXCEL_CONFIG dai valori di config.yaml."""
    excel_cfg = cfg.get("excel", {})
    if not excel_cfg:
        return
    mapping = {
        "header_font": "header_font",
        "header_size": "header_size",
        "header_bg": "header_bg",
        "header_fg": "header_fg",
        "data_font": "data_font",
        "data_size_download": "data_size",
        "alt_row_bg": "alt_row_bg",
        "border_color": "border_color",
        "max_col_width": "max_col_width",
    }
    for yaml_key, config_key in mapping.items():
        if yaml_key in excel_cfg:
            val = excel_cfg[yaml_key]
            # Rimuovi '#' iniziale dai colori se presente
            if isinstance(val, str) and val.startswith("#"):
                val = val[1:]
            EXCEL_CONFIG[config_key] = val


def create_excel(header, all_rows, output_path, sheet_name="Dati"):
    ec = EXCEL_CONFIG
    wb = Workbook()
    ws = wb.active
    ws.title = str(sheet_name)[:31]  # Excel limita a 31 caratteri

    header_font = Font(name=ec["header_font"], bold=True, size=ec["header_size"], color=ec["header_fg"])
    header_fill = PatternFill("solid", fgColor=ec["header_bg"])
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name=ec["data_font"], size=ec["data_size"])
    data_align = Alignment(vertical="center")
    thin_border = Border(
        left=Side(style="thin", color=ec["border_color"]),
        right=Side(style="thin", color=ec["border_color"]),
        top=Side(style="thin", color=ec["border_color"]),
        bottom=Side(style="thin", color=ec["border_color"]),
    )

    for col_idx, col_name in enumerate(header, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, row_data in enumerate(all_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            if col_idx <= len(header):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = data_align

    for col_idx in range(1, len(header) + 1):
        max_len = len(str(header[col_idx - 1]))
        for row in ws.iter_rows(min_row=2, max_row=min(50, ws.max_row), min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, ec["max_col_width"])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{ws.max_row}"

    wb.save(output_path)


def process_folder_pdf(folder_path, year):
    """
    Processa i PDF di una cartella e genera un Excel.
    Usa openpyxl in write-only mode + il generatore extract_data_from_pdf
    per tenere basso il consumo di RAM anche con PDF da milioni di righe.
    """
    import gc
    from openpyxl import Workbook as _Wb
    from openpyxl.cell.cell import WriteOnlyCell
    from openpyxl.styles import Font as _Fnt, PatternFill as _PF, Alignment as _Al

    pdf_files = sorted(glob.glob(os.path.join(folder_path, "*.pdf")))
    if not pdf_files:
        return False

    logging.info(f"  Trovati {len(pdf_files)} file PDF")

    header, signature, num_cols = None, None, 0
    for pdf_path in pdf_files:
        header, signature, num_cols = extract_header_from_pdf(pdf_path)
        if header:
            logging.info(f"  Intestazione ({num_cols} colonne): {header[:5]}...")
            break

    if not header:
        logging.error(f"  Impossibile determinare l'intestazione dai PDF")
        return False

    output_path = os.path.join(folder_path, f"dati_{year}.xlsx")

    # Write-only workbook: non carica nulla in RAM, scrive riga per riga
    ec = EXCEL_CONFIG
    wb = _Wb(write_only=True)
    ws = wb.create_sheet(title=str(year)[:31])

    # Riga di intestazione
    hdr_cells = []
    for col_name in header:
        c = WriteOnlyCell(ws, value=col_name)
        c.font      = _Fnt(name=ec["header_font"], bold=True, size=ec["header_size"], color=ec["header_fg"])
        c.fill      = _PF("solid", fgColor=ec["header_bg"])
        c.alignment = _Al(horizontal="center", vertical="center", wrap_text=True)
        hdr_cells.append(c)
    ws.append(hdr_cells)

    total_rows = 0
    skipped_cf = 0

    try:
        for pdf_path in pdf_files:
            pdf_name  = os.path.basename(pdf_path)
            file_rows = 0
            for row in extract_data_from_pdf(pdf_path, signature, num_cols):
                if not row[0].strip():      # salta righe senza codice fiscale
                    skipped_cf += 1
                    continue
                ws.append(row)
                file_rows  += 1
                total_rows += 1
            logging.info(f"  {pdf_name}: {file_rows} righe")
            gc.collect()

        if total_rows == 0:
            logging.warning(f"  Nessun dato estratto dai PDF")
            wb.close()
            return False

        if skipped_cf:
            logging.info(f"  Rimosse {skipped_cf} righe con codice fiscale vuoto")

        wb.save(output_path)
        logging.info(f"  => Creato: dati_{year}.xlsx ({total_rows} righe)")
        return True

    except Exception as e:
        logging.error(f"  Errore durante la creazione di dati_{year}.xlsx: {e}")
        if os.path.exists(output_path):
            os.remove(output_path)
        return False
    finally:
        wb.close()
        gc.collect()


def process_folder_csv(folder_path, year):
    """Processa i CSV di una cartella e genera un Excel."""
    csv_files = sorted(glob.glob(os.path.join(folder_path, "*.csv")))
    if not csv_files:
        return False

    logging.info(f"  Trovati {len(csv_files)} file CSV")

    all_rows = []
    master_header = None

    for csv_path in csv_files:
        csv_name = os.path.basename(csv_path)
        header, rows = extract_data_from_csv(csv_path)

        if header and master_header is None:
            master_header = header
            logging.info(f"  Intestazione ({len(header)} colonne): {header[:5]}...")

        # Normalizza il numero di colonne
        if master_header:
            target_cols = len(master_header)
            for row in rows:
                while len(row) < target_cols:
                    row.append("")
                del row[target_cols:]

        logging.info(f"  {csv_name}: {len(rows)} righe")
        all_rows.extend(rows)

    if not master_header or not all_rows:
        logging.warning(f"  Nessun dato estratto dai CSV")
        return False

    # Filtra righe con codice fiscale vuoto
    before = len(all_rows)
    all_rows = [r for r in all_rows if r[0].strip()]
    dropped = before - len(all_rows)
    if dropped:
        logging.info(f"  Rimosse {dropped} righe con codice fiscale vuoto")

    output_path = os.path.join(folder_path, f"dati_{year}.xlsx")
    create_excel(master_header, all_rows, output_path, year)
    logging.info(f"  => Creato: dati_{year}.xlsx ({len(all_rows)} righe)")
    return True

# ============================================================================
# INTERFACCIA INTERATTIVA
# ============================================================================

def ask_years(available_years, label="Aggiorna"):
    """Chiede quali anni elaborare."""
    print(f"\nAnni disponibili: {min(available_years)}-{max(available_years)}")
    print(f"  [1] {label} tutti gli anni")
    print(f"  [2] Scegli quali anni")
    print()

    while True:
        choice = input("Scelta [1/2]: ").strip()
        if choice == "1":
            return sorted(available_years, reverse=True)
        if choice == "2":
            while True:
                raw = input("Inserisci gli anni separati da virgola (es. 2020,2021,2022): ").strip()
                try:
                    years = [int(y.strip()) for y in raw.split(",") if y.strip()]
                    invalid = [y for y in years if y not in available_years]
                    if invalid:
                        print(f"  Anni non validi: {invalid}. Riprova.")
                        continue
                    if not years:
                        print("  Nessun anno inserito. Riprova.")
                        continue
                    return sorted(years, reverse=True)
                except ValueError:
                    print("  Formato non valido. Usa numeri separati da virgola.")
        print("  Scelta non valida, inserisci 1 o 2.")


def ask_source(year, has_pdf, has_csv):
    """Chiede se usare PDF o CSV per una cartella con entrambi."""
    print(f"\n  [{year}] Sono presenti sia file PDF che CSV.")
    print("    [1] Usa i CSV (piu' veloce, ma senza importi totali)")
    print("    [2] Usa i PDF (piu' lento, dati completi)")
    while True:
        choice = input("    Scelta [1/2]: ").strip()
        if choice == "1":
            return "csv"
        if choice == "2":
            return "pdf"
        print("    Scelta non valida.")


# ============================================================================
# MAIN
# ============================================================================

def parse_args():
    """Parsing degli argomenti da riga di comando."""
    parser = argparse.ArgumentParser(
        description="5 per Mille - Download e conversione dati dall'Agenzia delle Entrate",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""\
Esempi:
  %(prog)s                                    # modalita' interattiva
  %(prog)s --no-download --source csv         # solo conversione CSV, tutti gli anni
  %(prog)s --source csv --anni 2023,2024      # download + conversione CSV, anni scelti
  %(prog)s --no-download --no-convert         # non fa nulla (utile per test)
""",
    )
    parser.add_argument(
        "--root", type=str, default=".",
        help="Cartella root del progetto (default: directory corrente)",
    )
    parser.add_argument(
        "--anni", type=str, default=None,
        help="Anni da elaborare, separati da virgola (es. 2020,2021,2024). "
             "Se omesso, chiede interattivamente.",
    )
    parser.add_argument(
        "--source", choices=["pdf", "csv", "ask"], default="ask",
        help="Fonte dati per la conversione Excel: 'pdf', 'csv' o 'ask' "
             "(chiede per ogni cartella con entrambi i formati). Default: ask",
    )
    parser.add_argument(
        "--no-download", action="store_true",
        help="Salta la fase di download dall'Agenzia delle Entrate",
    )
    parser.add_argument(
        "--no-convert", action="store_true",
        help="Salta la fase di conversione in Excel",
    )
    parser.add_argument(
        "--yes", action="store_true",
        help="Risponde automaticamente Sì a tutte le domande interattive "
             "(download, conversione). Utile per uso da pipeline o cron.",
    )
    return parser.parse_args()


def main():
    args = parse_args()

    root_dir = os.path.abspath(args.root)
    os.makedirs(root_dir, exist_ok=True)

    # Carica configurazione esterna (se presente)
    cfg = load_config(root_dir)
    _apply_config(cfg)
    _apply_excel_config(cfg)

    get_logger("cinque_per_mille", root_dir)  # configura root logger
    logging.info(f"Directory root: {root_dir}")
    logging.info(f"Python: {sys.version}")
    if cfg:
        logging.info("Configurazione caricata da config.yaml")

    available_years = sorted(YEAR_URLS.keys(), reverse=True)

    # --- Parsing anni da CLI ---
    anni_cli = None
    if args.anni:
        try:
            anni_cli = sorted((int(a.strip()) for a in args.anni.split(",") if a.strip()), reverse=True)
            invalid = [a for a in anni_cli if a not in available_years]
            if invalid:
                logging.error(f"Anni non validi: {invalid}. Disponibili: {min(available_years)}-{max(available_years)}")
                sys.exit(1)
        except ValueError:
            logging.error(f"Formato anni non valido: '{args.anni}'. Usa numeri separati da virgola.")
            sys.exit(1)

    # ---- FASE 1: DOWNLOAD ----
    print("\n" + "=" * 60)
    print("  5 PER MILLE - Download e conversione dati")
    print("=" * 60)

    if args.no_download:
        logging.info("Download saltato (--no-download)")
        do_download = False
    elif not HAS_WEB:
        print("\nLe librerie 'requests' e 'beautifulsoup4' non sono installate.")
        print("   Il download non e' disponibile. Installa con:")
        print("   pip install requests beautifulsoup4")
        print("   Procedo solo con la conversione dei file gia' presenti.\n")
        do_download = False
    elif args.yes:
        logging.info("Download confermato automaticamente (--yes)")
        do_download = True
    else:
        do_download = ask_yes_no("\nVuoi aggiornare i file scaricandoli dal sito dell'Agenzia delle Entrate?")

    if do_download:
        years_to_download = anni_cli if anni_cli else ask_years(available_years, label="Scarica")
        logging.info(f"Anni da scaricare: {years_to_download}")

        session = requests.Session()
        total_files = 0

        for year in years_to_download:
            count = download_year(year, root_dir, session)
            total_files += count

        session.close()
        logging.info(f"\nDownload completato: {total_files} file totali")

    # ---- FASE 2: CONVERSIONE EXCEL ----
    if args.no_convert:
        logging.info("Conversione Excel saltata (--no-convert)")
        print(f"\nLog completo: {log_file}")
        return

    if not args.no_download and not do_download:
        # Modalita' interattiva: chiedi se procedere alla conversione
        pass  # prosegui comunque alla conversione
    if not args.no_convert:
        # In modalita' interattiva (no flag --no-download e --no-convert), chiedi conferma
        if not anni_cli and args.source == "ask":
            if not ask_yes_no("\nVuoi procedere alla creazione dei file Excel?"):
                logging.info("Conversione Excel saltata dall'utente.")
                print(f"\nLog completo: {log_file}")
                return

    # Trova le cartelle che hanno effettivamente file da elaborare
    years_with_data = []
    for year in available_years:
        folder = os.path.join(root_dir, str(year))
        if not os.path.isdir(folder):
            continue
        pdf_files = glob.glob(os.path.join(folder, "*.pdf"))
        csv_files = glob.glob(os.path.join(folder, "*.csv"))
        if pdf_files or csv_files:
            years_with_data.append(year)

    if not years_with_data:
        logging.warning("Nessuna cartella con dati da elaborare.")
        print(f"\nLog completo: {log_file}")
        return

    # Determina anni da elaborare
    if anni_cli:
        years_to_extract = [y for y in anni_cli if y in years_with_data]
        if not years_to_extract:
            logging.warning(f"Nessuna cartella con dati per gli anni richiesti: {anni_cli}")
            print(f"\nLog completo: {log_file}")
            return
    else:
        years_to_extract = ask_years(years_with_data, label="Elabora")

    logging.info(f"Anni da elaborare: {years_to_extract}")

    print()
    processed = 0
    errors = 0

    for year in years_to_extract:
        folder = os.path.join(root_dir, str(year))
        if not os.path.isdir(folder):
            continue

        pdf_files = glob.glob(os.path.join(folder, "*.pdf"))
        csv_files = glob.glob(os.path.join(folder, "*.csv"))

        if not pdf_files and not csv_files:
            logging.info(f"[{year}] Cartella vuota, salto")
            continue

        logging.info(f"[{year}] Elaborazione ({len(pdf_files)} PDF, {len(csv_files)} CSV)")

        # Scegli la fonte dati
        has_pdf = len(pdf_files) > 0
        has_csv = len(csv_files) > 0

        if args.source in ("pdf", "csv"):
            # Fonte specificata da CLI
            source = args.source
            if source == "pdf" and not has_pdf:
                logging.warning(f"[{year}] Nessun PDF disponibile, uso CSV")
                source = "csv"
            elif source == "csv" and not has_csv:
                logging.warning(f"[{year}] Nessun CSV disponibile, uso PDF")
                source = "pdf"
        elif has_pdf and has_csv:
            source = ask_source(year, has_pdf, has_csv)
        elif has_csv:
            source = "csv"
        else:
            source = "pdf"

        logging.info(f"[{year}] Fonte dati: {source.upper()}")

        try:
            if source == "csv":
                success = process_folder_csv(folder, year)
            else:
                success = process_folder_pdf(folder, year)

            if success:
                processed += 1
                # Copia dati_{year}.xlsx in Dati/ per etl.py
                dati_dir = os.path.join(root_dir, "Dati")
                os.makedirs(dati_dir, exist_ok=True)
                src = os.path.join(folder, f"dati_{year}.xlsx")
                if os.path.isfile(src):
                    shutil.copy2(src, os.path.join(dati_dir, f"dati_{year}.xlsx"))
                    logging.info(f"[{year}] Copiato in Dati/dati_{year}.xlsx")
            else:
                errors += 1
        except Exception as e:
            logging.error(f"[{year}] Errore imprevisto: {e}")
            errors += 1

    print("\n" + "=" * 60)
    logging.info(f"Completato: {processed} cartelle elaborate, {errors} errori")
    print(f"Log completo: {log_file}")
    print("=" * 60)


if __name__ == "__main__":
    main()
