#!/usr/bin/env python3
"""
Converte tutti i PDF contenuti nelle cartelle annuali (2006-2025) in un unico file Excel per cartella.

Uso:
    python pdf_to_excel.py [percorso_root]

Se non viene specificato un percorso, usa la directory corrente.
Le cartelle devono chiamarsi con l'anno (es. "2006", "2007", ..., "2025").
Lo script cerca tutti i file .pdf in ogni cartella, li unisce e produce un file .xlsx nella stessa cartella.

Requisiti:
    pip install pdfplumber openpyxl
"""

import sys
import os
import re
import glob
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def clean_cell(value):
    """Pulisce il contenuto di una cella: rimuove newline interni e spazi multipli."""
    if value is None:
        return ""
    value = str(value).strip()
    value = re.sub(r'\n+', ' ', value)
    value = re.sub(r'\s{2,}', ' ', value)
    # Rimuove caratteri di controllo illegali per Excel (tranne tab e newline gia' gestiti)
    value = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', value)
    return value


def extract_header(tables_first_page):
    """
    Estrae l'intestazione dalla prima tabella della prima pagina.
    Gestisce intestazioni multi-riga (tipicamente 2 righe con merge).
    Restituisce una lista di stringhe, una per colonna.
    """
    if not tables_first_page:
        return None

    table = tables_first_page[0]
    if len(table) < 2:
        return [clean_cell(c) for c in table[0]]

    row1 = table[0]
    row2 = table[1]
    num_cols = max(len(row1), len(row2))

    # Verifica se la seconda riga e' effettivamente una sotto-intestazione
    # (cioe' contiene None nelle prime colonne, tipico di un merge)
    is_two_row_header = row2[0] is None

    if not is_two_row_header:
        return [clean_cell(c) for c in row1]

    # Costruisce intestazione combinata dalle due righe
    header = []
    for i in range(num_cols):
        top = clean_cell(row1[i]) if i < len(row1) else ""
        bottom = clean_cell(row2[i]) if i < len(row2) else ""

        if top and bottom:
            header.append(f"{top} - {bottom}")
        elif top:
            header.append(top)
        elif bottom:
            header.append(bottom)
        else:
            header.append(f"Colonna_{i+1}")

    return header


def is_header_row(row, header_signature):
    """
    Verifica se una riga e' un'intestazione ripetuta confrontandola
    con la 'firma' dell'intestazione originale (le prime celle non vuote).
    """
    cleaned = [clean_cell(c) for c in row]
    # Confronta le prime N celle non vuote
    row_sig = [c for c in cleaned if c][:3]
    return row_sig == header_signature


def extract_data_from_pdf(pdf_path, header_signature, num_cols):
    """
    Estrae tutte le righe dati da un PDF, saltando le intestazioni ripetute.
    Restituisce una lista di liste (righe x colonne).
    """
    rows = []
    try:
        pdf = pdfplumber.open(pdf_path)
    except Exception as e:
        print(f"  ERRORE nell'aprire {pdf_path}: {e}")
        return rows

    for page_idx, page in enumerate(pdf.pages):
        try:
            tables = page.extract_tables()
        except Exception as e:
            print(f"  ERRORE pagina {page_idx + 1} di {os.path.basename(pdf_path)}: {e}")
            continue

        for table in tables:
            for row in table:
                cleaned = [clean_cell(c) for c in row]

                # Salta righe che sono intestazioni ripetute
                if is_header_row(row, header_signature):
                    continue

                # Salta righe completamente vuote
                if all(c == "" or c is None for c in cleaned):
                    continue

                # Salta sotto-intestazioni (seconda riga dell'header, inizia con None/vuoto)
                if cleaned[0] == "" and any(c != "" for c in cleaned[4:]):
                    non_empty = [c for c in cleaned if c]
                    # Euristica: se i valori non vuoti sono tutti brevi e sembrano
                    # etichette di sotto-colonna, e' una sotto-intestazione
                    if all(len(c) < 40 for c in non_empty) and len(non_empty) <= num_cols // 2:
                        continue

                # Normalizza il numero di colonne
                while len(cleaned) < num_cols:
                    cleaned.append("")
                cleaned = cleaned[:num_cols]

                rows.append(cleaned)

    pdf.close()
    return rows


def create_excel(header, all_rows, output_path, year):
    """Crea il file Excel con formattazione professionale."""
    wb = Workbook()
    ws = wb.active
    ws.title = str(year)

    # Stili
    header_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # Intestazione
    for col_idx, col_name in enumerate(header, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Dati
    for row_idx, row_data in enumerate(all_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # Auto-dimensiona colonne (approssimativo)
    for col_idx in range(1, len(header) + 1):
        max_len = len(str(header[col_idx - 1]))
        for row in ws.iter_rows(min_row=2, max_row=min(50, ws.max_row), min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 45)

    # Freeze panes e filtro
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{ws.max_row}"

    wb.save(output_path)


def process_folder(folder_path, year):
    """Processa tutti i PDF in una cartella e genera un unico Excel."""
    pdf_files = sorted(glob.glob(os.path.join(folder_path, "*.pdf")))

    if not pdf_files:
        print(f"  Nessun PDF trovato in {folder_path}")
        return False

    print(f"  Trovati {len(pdf_files)} file PDF")

    # Fase 1: determina l'intestazione dal primo PDF
    header = None
    header_signature = None
    num_cols = 0

    for pdf_path in pdf_files:
        try:
            pdf = pdfplumber.open(pdf_path)
            first_page_tables = pdf.pages[0].extract_tables()
            pdf.close()

            header = extract_header(first_page_tables)
            if header:
                num_cols = len(header)
                # Firma dell'intestazione: le prime 3 celle non vuote della prima riga
                first_row = [clean_cell(c) for c in first_page_tables[0][0]]
                header_signature = [c for c in first_row if c][:3]
                print(f"  Intestazione rilevata ({num_cols} colonne): {header[:5]}...")
                break
        except Exception as e:
            print(f"  Errore nel leggere intestazione da {os.path.basename(pdf_path)}: {e}")
            continue

    if not header:
        print(f"  ERRORE: impossibile determinare l'intestazione per {folder_path}")
        return False

    # Fase 2: estrai dati da tutti i PDF
    all_rows = []
    for pdf_path in pdf_files:
        pdf_name = os.path.basename(pdf_path)
        print(f"  Elaboro: {pdf_name}...", end=" ")
        rows = extract_data_from_pdf(pdf_path, header_signature, num_cols)
        print(f"{len(rows)} righe")
        all_rows.extend(rows)

    if not all_rows:
        print(f"  ATTENZIONE: nessun dato estratto per {folder_path}")
        return False

    # Fase 3: crea l'Excel
    output_name = f"dati_{year}.xlsx"
    output_path = os.path.join(folder_path, output_name)
    create_excel(header, all_rows, output_path, year)
    print(f"  => Creato: {output_name} ({len(all_rows)} righe totali)")
    return True


def main():
    root_dir = sys.argv[1] if len(sys.argv) > 1 else "."
    root_dir = os.path.abspath(root_dir)

    print(f"Directory root: {root_dir}")
    print(f"Cerco cartelle dal 2006 al 2025...\n")

    found = 0
    processed = 0

    for year in range(2006, 2026):
        folder_path = os.path.join(root_dir, str(year))
        if os.path.isdir(folder_path):
            found += 1
            print(f"[{year}] Cartella trovata")
            success = process_folder(folder_path, year)
            if success:
                processed += 1
            print()

    if found == 0:
        print("Nessuna cartella annuale trovata (cercavo 2006/, 2007/, ..., 2025/).")
        print(f"Verifica che le cartelle siano presenti in: {root_dir}")
        sys.exit(1)

    print(f"Completato: {processed}/{found} cartelle elaborate con successo.")


if __name__ == "__main__":
    main()
