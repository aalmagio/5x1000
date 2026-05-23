#!/usr/bin/env python3
"""
estrai_scelte_categorie.py
==========================
Estrae dalle ultime righe dei file xlsx di categoria il numero di:
  - scelte espresse (ammessi)
  - scelte generiche (ammessi)
  - scelte espresse (esclusi)

per ogni categoria e anno (2019 in avanti) e produce un Excel pivot.

Uso:
    python estrai_scelte_categorie.py
    python estrai_scelte_categorie.py --dati Dati --output scelte_categorie.xlsx
"""

import argparse
import logging
import warnings
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Ordine e label display delle categorie
# ---------------------------------------------------------------------------
CAT_ORDER = [
    ("Ricerca_scientifica", "Ricerca scientifica"),
    ("Ricerca_sanitaria",   "Ricerca sanitaria"),
    ("ETS_ONLUS",           "ETS ONLUS"),
    ("Comuni",              "Comuni"),
    ("Beni_culturali",      "Beni culturali"),
    ("ASD",                 "ASD"),
    ("Aree_protette",       "Aree protette"),
]


def _parse_num_it(val) -> int:
    """Converte stringa italiana (es. '542.053') in intero."""
    if val is None:
        return 0
    s = str(val).strip().replace(".", "").replace(",", "").replace(" ", "")
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return 0


# Alias per colonna N_SCELTE (stessi di db_updater.py)
_N_SCELTE_ALIASES = frozenset({
    "numero scelte", "n. scelte", "n scelte", "numero di scelte",
    "preferenze", "n preferenze", "numero preferenze",
})


def _trova_col_scelte(rows) -> int | None:
    """
    Scansiona le prime 5 righe cercando l'intestazione della colonna N_SCELTE.
    Ritorna l'indice di colonna (0-based) oppure None se non trovata.
    """
    for row in rows[:5]:
        header = [str(c.value or "").strip().lower() for c in row]
        # Match esatto
        for i, h in enumerate(header):
            if h in _N_SCELTE_ALIASES:
                return i
        # Match parziale (alias lunghi)
        for i, h in enumerate(header):
            for a in _N_SCELTE_ALIASES:
                if len(a) > 6 and (a in h or h in a):
                    return i
    return None


def _val_footer(row, n_col: int | None):
    """
    Ritorna il valore dalla colonna n_col (se disponibile e non None),
    altrimenti il primo valore non-None dalla colonna 1 in poi.
    """
    if n_col is not None and n_col < len(row) and row[n_col].value is not None:
        return row[n_col].value
    for cell in row[1:]:
        if cell.value is not None:
            return cell.value
    return None


def _leggi_footer(path: Path) -> tuple[int, int]:
    """
    Legge un file xlsx di categoria e restituisce (scelte_espresse, scelte_generiche).

    Prima cerca footer rows (presenti nei file estratti da PDF):
      - "TOTALE SCELTE ESPRESSE" → scelte espresse
      - "Voti generici"          → scelte generiche
      - "TOTALE" (solo se nessuna riga espresse trovata, caso Comuni) → scelte espresse
      - "SUB TOTALE" (schema 2019) → scelte espresse

    Se non trova footer rows (file estratti da CSV senza righe di totale),
    somma direttamente la colonna N_SCELTE da tutti i dati.
    """
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    ws = wb.active
    # values_only=False necessario: alcune celle footer restituiscono None con values_only=True
    rows = list(ws.iter_rows(values_only=False))
    wb.close()

    n_col = _trova_col_scelte(rows)

    espresse  = 0
    generiche = 0
    found_esp = False

    # Scansiona le ultime 20 righe (i footer sono sempre in fondo)
    for row in rows[-20:]:
        if not row or row[0].value is None:
            continue
        label = str(row[0].value).strip().lower()
        val   = _val_footer(row, n_col)
        logger.debug(f"  footer scan [{path.name}] label={label!r:50s} val={val!r}")

        if "totale scelte espresse" in label:
            espresse  = _parse_num_it(val)
            found_esp = True
        elif label == "sub totale" and not found_esp:
            # Schema 2019: "SUB TOTALE" contiene le scelte espresse
            espresse  = _parse_num_it(val)
            found_esp = True
        elif "generici" in label or "generiche" in label or "generico" in label:
            generiche = _parse_num_it(val)
        elif label == "totale" and not found_esp:
            # Schema Comuni: TOTALE contiene le scelte espresse
            espresse  = _parse_num_it(val)
            found_esp = True

    # Fallback: file da CSV senza footer rows → somma colonna N_SCELTE dai dati
    if not found_esp and n_col is not None:
        total = 0
        for row in rows[1:]:  # salta l'header
            if not row or row[0].value is None:
                continue
            label = str(row[0].value).strip().lower()
            # salta eventuali righe footer (non dati)
            if any(kw in label for kw in ("totale", "sub totale", "generici", "generiche")):
                continue
            v = row[n_col].value if n_col < len(row) else None
            total += _parse_num_it(v)
        if total > 0:
            espresse = total

    return espresse, generiche


def estrai_dati(dati_dir: Path, anno_min: int = 2019) -> dict:
    """
    Scansiona le sottocartelle di categoria in dati_dir e raccoglie
    per ogni (categoria, anno): scelte_esp_amm, scelte_gen_amm, scelte_esp_esc.

    Struttura result:
      result[cat_key][anno] = {
          "esp_amm": int,
          "gen_amm": int,
          "esp_esc": int,
      }
    """
    cat_keys = {c[0].upper(): c[0] for c in CAT_ORDER}
    result: dict[str, dict[int, dict]] = {c[0]: {} for c in CAT_ORDER}

    anni_trovati: set[int] = set()

    for cat_dir in dati_dir.iterdir():
        if not cat_dir.is_dir() or cat_dir.name.startswith("."):
            continue

        # Mappa nome cartella → chiave canonica
        cat_key = cat_keys.get(cat_dir.name.upper())
        if cat_key is None:
            logger.debug(f"Cartella ignorata: {cat_dir.name}")
            continue

        for stato, pattern in [("ammesso", "*ammessi*.xlsx"), ("escluso", "*esclusi*.xlsx")]:
            for xlsx in sorted(cat_dir.glob(pattern)):
                try:
                    anno = int(xlsx.stem.split("_")[0])
                except (IndexError, ValueError):
                    logger.warning(f"Anno non estraibile da {xlsx.name}")
                    continue
                if anno < anno_min:
                    continue

                anni_trovati.add(anno)
                esp, gen = _leggi_footer(xlsx)
                logger.info(
                    f"  {cat_key}/{anno}/{stato}: "
                    f"espresse={esp:,}  generiche={gen:,}"
                )

                if anno not in result[cat_key]:
                    result[cat_key][anno] = {"esp_amm": 0, "gen_amm": 0, "esp_esc": 0}

                if stato == "ammesso":
                    result[cat_key][anno]["esp_amm"] = esp
                    result[cat_key][anno]["gen_amm"] = gen
                else:
                    result[cat_key][anno]["esp_esc"] = esp

    return result, sorted(anni_trovati)


def scrivi_excel(result: dict, anni: list[int], output: Path) -> None:
    """
    Scrive il workbook Excel pivot con la struttura:
      Categoria | stato | tipo | anno1 | anno2 | ...
    """
    wb   = openpyxl.Workbook()
    ws   = wb.active
    ws.title = "Scelte per categoria"

    # ── Stili ──────────────────────────────────────────────────────────────
    HEADER_FILL  = PatternFill("solid", fgColor="2E5B9A")
    CAT_FILL     = PatternFill("solid", fgColor="D9E1F2")
    AMM_FILL_1   = PatternFill("solid", fgColor="EBF3EC")
    AMM_FILL_2   = PatternFill("solid", fgColor="FFFFFF")
    ESC_FILL     = PatternFill("solid", fgColor="FFF2CC")
    TOTAL_FILL   = PatternFill("solid", fgColor="FCE4D6")

    HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
    LABEL_FONT   = Font(bold=True, size=10)
    NUM_FONT     = Font(size=10)
    TOTAL_FONT   = Font(bold=True, size=10)

    thin  = Side(style="thin",   color="BFBFBF")
    thick = Side(style="medium", color="7F7F7F")
    BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)
    THICK_BORD  = Border(left=thick, right=thick, top=thick, bottom=thick)

    def _cell(row, col, value, font=None, fill=None, align=None, num_fmt=None):
        c = ws.cell(row=row, column=col, value=value)
        if font:    c.font      = font
        if fill:    c.fill      = fill
        if align:   c.alignment = align
        if num_fmt: c.number_format = num_fmt
        c.border = BORDER
        return c

    center = Alignment(horizontal="center", vertical="center")
    right  = Alignment(horizontal="right",  vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    NUM_FMT = '#,##0'

    # ── Header ─────────────────────────────────────────────────────────────
    headers = ["Categoria", "stato", "tipo"] + [str(a) for a in anni]
    for col, h in enumerate(headers, 1):
        _cell(1, col, h, font=HEADER_FONT, fill=HEADER_FILL, align=center)

    # ── Righe dati ─────────────────────────────────────────────────────────
    row_idx = 2
    tot_esp = {a: 0 for a in anni}
    tot_gen = {a: 0 for a in anni}
    tot_esc = {a: 0 for a in anni}

    for cat_key, cat_label in CAT_ORDER:
        cat_data = result.get(cat_key, {})

        rows_cat = [
            ("ammessi", "scelta",   "esp_amm", AMM_FILL_1),
            ("ammessi", "generica", "gen_amm", AMM_FILL_2),
            ("esclusi", "scelta",   "esp_esc", ESC_FILL),
        ]

        for stato, tipo, field, fill in rows_cat:
            _cell(row_idx, 1, cat_label, font=LABEL_FONT, fill=CAT_FILL,  align=left)
            _cell(row_idx, 2, stato,     font=NUM_FONT,   fill=fill, align=center)
            _cell(row_idx, 3, tipo,      font=NUM_FONT,   fill=fill, align=center)

            for col, anno in enumerate(anni, 4):
                val = cat_data.get(anno, {}).get(field, 0)
                _cell(row_idx, col, val if val else None,
                      font=NUM_FONT, fill=fill, align=right, num_fmt=NUM_FMT)
                if val:
                    if field == "esp_amm": tot_esp[anno] += val
                    if field == "gen_amm": tot_gen[anno] += val
                    if field == "esp_esc": tot_esc[anno] += val

            row_idx += 1

    # ── Righe totali ────────────────────────────────────────────────────────
    TOTAL_FILL2 = PatternFill("solid", fgColor="F2DCDB")
    TOTAL_FILL3 = PatternFill("solid", fgColor="E2EFDA")
    GRAND_FILL  = PatternFill("solid", fgColor="D6DCE4")
    YOY_FILL    = PatternFill("solid", fgColor="EDEDED")

    def _total_row(label, src, fill, fmt=NUM_FMT):
        nonlocal row_idx
        _cell(row_idx, 1, label, font=TOTAL_FONT, fill=fill, align=left)
        _cell(row_idx, 2, None,  font=TOTAL_FONT, fill=fill)
        _cell(row_idx, 3, None,  font=TOTAL_FONT, fill=fill)
        for col, anno in enumerate(anni, 4):
            val = src.get(anno, 0)
            _cell(row_idx, col, val if val else None,
                  font=TOTAL_FONT, fill=fill, align=right, num_fmt=fmt)
        row_idx += 1

    _total_row("TOTALE scelte espresse ammessi", tot_esp, TOTAL_FILL)
    _total_row("TOTALE scelte generiche",        tot_gen, TOTAL_FILL2)
    _total_row("TOTALE scelte esclusi",          tot_esc, TOTAL_FILL3)

    tot_valide = {a: tot_esp[a] + tot_gen[a] for a in anni}
    tot_firme  = {a: tot_esp[a] + tot_gen[a] + tot_esc[a] for a in anni}
    _total_row("Totale firme valide", tot_valide, GRAND_FILL)
    _total_row("Totale firme",        tot_firme,  GRAND_FILL)

    # Crescita YoY su Totale firme
    _cell(row_idx, 1, "Crescita YoY", font=TOTAL_FONT, fill=YOY_FILL, align=left)
    _cell(row_idx, 2, None, font=TOTAL_FONT, fill=YOY_FILL)
    _cell(row_idx, 3, None, font=TOTAL_FONT, fill=YOY_FILL)
    for col, anno in enumerate(anni, 4):
        idx = anni.index(anno)
        prev = tot_firme.get(anni[idx - 1], 0) if idx > 0 else 0
        curr = tot_firme.get(anno, 0)
        yoy  = (curr - prev) / prev if (idx > 0 and prev) else None
        _cell(row_idx, col, yoy, font=TOTAL_FONT, fill=YOY_FILL,
              align=right, num_fmt="0.0%")
    row_idx += 1

    # ── Larghezze colonne ───────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    for col in range(4, 4 + len(anni)):
        ws.column_dimensions[get_column_letter(col)].width = 13

    # Blocca intestazione
    ws.freeze_panes = "D2"

    wb.save(output)
    logger.info(f"Salvato: {output}")


def main():
    ap = argparse.ArgumentParser(description="Estrai scelte per categoria da file xlsx AdE")
    ap.add_argument("--dati",   default="Dati",
                    help="Cartella Dati/ (default: ./Dati)")
    ap.add_argument("--output", default="Dati/scelte_categorie.xlsx",
                    help="File Excel di output (default: Dati/scelte_categorie.xlsx)")
    ap.add_argument("--anno-min", type=int, default=2019,
                    help="Anno minimo (default: 2019)")
    args = ap.parse_args()

    dati_dir = Path(args.dati)
    if not dati_dir.exists():
        logger.error(f"Cartella non trovata: {dati_dir}")
        return

    logger.info("Scansione file categoria...")
    result, anni = estrai_dati(dati_dir, anno_min=args.anno_min)

    if not anni:
        logger.error("Nessun file trovato. Controlla il percorso Dati/.")
        return

    logger.info(f"Anni trovati: {anni}")
    scrivi_excel(result, anni, Path(args.output))


if __name__ == "__main__":
    main()
